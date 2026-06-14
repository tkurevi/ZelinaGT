"""
===============================================================================
economy_Ze1.py  -  Zelina-1 (Ze-1) GEOTHERMAL DOUBLET LOW-T / HP-ASSISTED DH ECONOMIC MODEL
===============================================================================
Master feasibility model.  Run from the SAME FOLDER as the V3 production model
(in Spyder: click play here).  V3 is the unchanged "physics master"; this script
runs alongside it and is FED from it.  It integrates four engines:

    V3 reservoir/well model ...  production operating point + mirror-well injection
    esp_geothermal.py .........  ESP + injection-pump electrical power
    geothermal_HE.py ..........  PHE sizing, buried-pipeline heat loss, delivered
                                 heat, DH circulation pump  (revenue = delivered)
    doublet_decline.py ........  Gringarten-Sauty thermal breakthrough T_prod(t)

and adds the full economics: CAPEX, OPEX, debt financing, escalation, tax/
depreciation, and metrics (NPV, IRR, simple & discounted payback, LCOH, DSCR)
plus a one-at-a-time tornado.  Console report + Excel workbook + plots.

ALL unit costs are EDITABLE ASSUMPTIONS (generic Croatian/EU feasibility values,
flagged below).  Replace with quotes when available.
===============================================================================
"""
import os, sys, math, io, contextlib
from dataclasses import dataclass, field, asdict
import numpy as np
import plot_style  # uniform figure style (applied on import)

# --- per-well thermal / gas constants (for Alves wellhead-T + Jelic props) ---
T_BH = 52.9      # bottomhole flowing (mixed inflow) temperature, C (static gradient @884 m)
GWR_M3M3 = 0.0    # gas-water ratio (Ze-1 ~ single-phase fresh water, UV-negative)
SUBBASIN = "Sava"
GEO_GRAD = 0.043
SALINITY = 870
WELL_NAME = "Ze-1"
HP_SPF_DEFAULT = 4.7   # seasonal performance factor of the consumer-side heat pumps


THIS_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, THIS_DIR)

# ===========================================================================
# 1) CONFIG  -  every input lives here
# ===========================================================================
@dataclass
class Config:
    # ---- project / financing -------------------------------------------------
    project_life_yr:        int   = 30
    discount_rate:          float = 0.05      # [-] real discount rate (input)
    debt_ratio:             float = 0.50      # [-] 0=all equity, 1=all debt
    loan_interest:          float = 0.035     # [-] HBOR/commercial energy loan
    loan_tenor_yr:          int   = 17        # [yr] https://www.hbor.hr/krediti-za-energetsku-ucinkovitost-poduzetnika/2451
    tax_rate:               float = 0.18      # [-] Croatian corporate tax
    inflation:              float = 0.03      # [-] general (real analysis -> 0)
    heat_price_escal:       float = 0.03      # [-] heat-price escalation (manual)
    elec_price_escal:       float = 0.03      # [-] electricity escalation (manual)
    opex_escal:             float = 0.03      # [-] opex escalation (manual)

    # ---- energy market -------------------------------------------------------
    heat_price_eur_MWhth:   float = 35.00     # [EUR/MWhth] heat selling price; no tax incl.
    elec_price_eur_MWhe:    float = 135.0     # [EUR/MWhe]  electricity (uniform) no tax incl.
    capacity_price_eur_kW_yr: float = 16.65   # [EUR/kWth/yr] DH capacity payment no tax incl.
    co2_price_eur_t:        float = 70.0      # [EUR/tCO2]
    co2_as_revenue:         bool  = False     # Y/N: count CO2 credit as revenue?
    hp_capex_in_scope:      bool  = True      # Y/N: include consumer/central heat-pump CAPEX
    hp_elec_in_opex:        bool  = True      # Y/N: include heat-pump electricity in OPEX
    gas_boiler_eff:         float = 0.85
    ng_emission_t_MWh:      float = 0.202     # tCO2/MWh_gas (LHV)

    # ---- operation -----------------------------------------------------------
    FLH:                    float = 2000.0    # [h/yr] annual full-load hours (heat)
    dh_scenario:            str   = "C"       # Zelina = HP-assisted ambient loop (15/10 C); see hp_spf

    # ---- geometry / distances (inputs) --------------------------------------
    city_distance_m:        float = 150.0    # Zelina: consumer 150 m from well (x2 supply+return)
    injection_distance_m:   float = 600.0     # producer(new) -> injector(Ze-1) = doublet spacing
    grid_distance_m:        float = 500.0     # to nearest MV connection point
    dh_pipe_DN_m:           float = 0.2101    # carrier ID (DN200)
    brine_pipe_DN_m:        float = 0.20      # injection line ID (DN200, low dP)

    # ---- doublet (MANUAL inputs - user controls) ----------------------------
    doublet_avg_flow_ls:    float = 25.9      # [L/s] design circulation flow (single operating point)
    operating_months_per_yr:int   = 12         # single operating point -> peak = avg = 25.9 L/s (editable for seasonal duty)
    doublet_spacing_m:      float = 600.0      # producer(new)<->injector(Ze-1) spacing (FEFLOW/own)
    doublet_method:         str   = "GS"
    barends_dispersivity_m: float = 5.0       # [m] Barends longitudinal dispersivity a_L (per layer); used only when doublet_method="GS+Barends"
    barends_underburden:    bool  = True       # Barends bleeding: True=cap+base (h_eff=2x), False=cap only; used only when doublet_method="GS+Barends"

    # ---- CAPEX unit costs (EDITABLE feasibility assumptions) -----------------
    well_cost_eur_per_m:    float = 3950.0    # turnkey EUR/m (new production well)
    prod_well_cost_eur:     float = 0.0       # 0 => compute new producer cost = well_cost_eur_per_m * prod_well_depth_m
    inj_well_depth_m:       float = 891.0     # Ze-1 reservoir base (EXISTING well; cost via inj_well_cost_eur)
    prod_well_depth_m:      float = 900.0     # NEW producer well drilled depth (9 m below S7 base @891 m)
    inj_well_cost_eur:      float = 60000.0   # Ze-1 -> injector conversion (workover+wellhead); well itself already paid
    hp_spf:                 float = 4.7       # consumer heat-pump SPF (Q_delivered = Q_geo*SPF/(SPF-1))
    hp_eur_per_kW:          float = 400.0     # heat-pump CAPEX per kW(thermal) delivered (set 0 if consumer-owned)
    dh_supply_C:            float = 15.0      # ambient-loop supply to consumer HPs
    dh_return_C:            float = 10.0      # ambient-loop return from consumer HPs
    he_cold_approach_K:     float = 3.0       # brine_out (injection T) = dh_return + cold approach
    esp_eur_per_kW:         float = 1200.0    # ESP pump+motor+tubing 4 1/2 (surface kW)
    esp_cable_eur_per_m:    float = 45.0      # downhole power cable
    esp_install_eur:        float = 120000.0  # service-company install (lump)
    injpump_eur_per_kW:     float = 700.0     # injection pump
    injpump_install_eur:    float = 40000.0
    phe_eur_per_m2:         float = 700.0     # plate heat exchanger; depended on inox/titanium options
    circ_eur_per_kW:        float = 600.0     # DH circulation pump
    plant_prod_eur:         float = 50000.0   # energy-plant container @ production well
    plant_inj_eur:          float = 30000.0   # container @ injection well
    dh_pipe_eur_per_m:      float = 1000.0    # buried pre-insulated DN200 (per metre of trench)+installation
    brine_pipe_eur_per_m:   float = 650.0     # buried brine line DN200 no insulation + installation
    transformer_eur_per_kVA:float = 50.0
    mv_line_eur_per_m:      float = 120.0
    grid_connection_fee_eur:float = 40000.0   # HERA 96.5 EUR/kW adjust
    eng_pct:                float = 0.03      # engineering/design/control projects as percentage of capex
    contingency_pct:        float = 0.10      # contingency

    # ---- OPEX (EDITABLE) -----------------------------------------------------
    personnel_eur_yr:       float = 30000.0   # 1 service person (Croatian mean gross + contrib.)
    grid_capacity_charge_eur_kW_yr: float = 10.0   # grid DEMAND charge (COST) on connected kW (flat annual fee on the peak or contracted connected power no matter using or not)
    sm_pct_surface:         float = 0.015     # service&maint, % of surface mech/elec capex
    sm_pct_wells:           float = 0.020     # service&maint, % of well capex
    chemicals_eur_yr:       float = 10000.0   # scaling/corrosion inhibitor + cleaning
    insurance_pct:          float = 0.002     # of total capex
    misc_opex_eur_yr:       float = 10000.0   # telemetry, land, adminstration
    esp_replace_interval_yr:int   = 8         # ESP replacement (capex event)
    injpump_replace_interval_yr: int = 12     # Inj. pump replacement (capex event)

    # ---- concession (Croatian, inputs) --------------------------------------
    field_area_km2:         float = 10.0
    concession_fixed_eur_km2: float = 132.72
    concession_var_pct:     float = 0.03      # 3% of (heat price * delivered kWh)

    # ---- injectivity --------------------------------------------------------
    injectivity_multiplier: float = 1.0       # 1.0 = mirror (=productivity); de-rate if known


# ===========================================================================
# 2) V3 INTERFACE  -  pull production point + mirror-well injection from V3
#    Tries to import the V3 reservoir model; otherwise uses the values already
#    derived from V3 for the 15 L/s design point (documented constants).
# ===========================================================================
def get_v3_results(target_flow_ls=25.9, reinj_T_C=13.0, cfg=None):
    """Ze-1 IVA 7-layer commingled reservoir feeding the economics (from main_ze1).
    PRODUCER = new well (9-5/8 to 753 m + slotted liner; ESP seatable ~700 m).
    INJECTOR = existing Ze-1 (9-5/8 to 268 m; cannot host the design-rate ESP).
    """
    # ---- validated 7-layer set (mirror of main_ze1.reservoir_layers) ----------
    layers = [dict(h=11.0, phi=0.322, k=93.9, T0=48.7),
              dict(h=13.0, phi=0.319, k=90.8, T0=49.4),
              dict(h=8.5,  phi=0.317, k=88.3, T0=49.9),
              dict(h=4.5,  phi=0.313, k=84.0, T0=50.6),
              dict(h=12.0, phi=0.309, k=79.3, T0=51.6),
              dict(h=15.0, phi=0.310, k=81.0, T0=52.1),
              dict(h=14.0, phi=0.297, k=68.4, T0=52.9)]
    res = dict(
        prod_flow_ls   = target_flow_ls,
        wellhead_T_C   = 48.0,             # produced WH T (warm at design rate); refined by VLP below
        Pwf_bar        = 59.7,             # flowing BHP @ z_ref(884 m) at 25.9 L/s
        static_bar     = 93.5,
        drawdown_bar   = 33.8,
        dynamic_level_m= 700.0,            # ESP seat (new well); pumping lift datum
        esp_depth_m    = 700.0,
        esp_intake_bar = 20.0,
        inj_bhp_bar    = 127.3,            # mirror: 2*static - Pwf
        inj_overpres_bar = 33.8,
        inj_depth_m    = 884.0,
        layers         = layers,
        source         = "Ze-1 fallback constants (25.9 L/s design, new-well producer)")
    try:
        from main_ze1 import build_commingled, NEWWELL_PRODUCER         # type: ignore
        from pvt import bar_to_Pa, Pa_to_bar, GRAVITY                    # type: ignore
        import numpy as np
        comm = build_commingled(NEWWELL_PRODUCER)
        res["inj_depth_m"] = comm.z_ref
        Pwf = Pa_to_bar(comm.Pwf_at_q(target_flow_ls/1000.0)); res["Pwf_bar"] = round(Pwf,1)
        ps = np.arange(55, 96, 2.0); qs = np.array([comm.q_at_Pwf(bar_to_Pa(p))*1000 for p in ps])
        m = qs > 0.5; slope, inter = np.polyfit(ps[m], qs[m], 1); static = -inter/slope
        res["static_bar"]   = round(static,1)
        res["drawdown_bar"] = round(static - Pwf,1)
        res["inj_bhp_bar"]  = round(2*static - Pwf,1)            # mirror well, injectivity = productivity
        res["inj_overpres_bar"] = round(static - Pwf,1)
        rho = getattr(comm, "rho_wb", 1000.0)
        # ESP for the NEW producer sits at 700 m (inside its 9-5/8" casing to 753 m)
        esp = 700.0
        res["esp_depth_m"]    = float(esp)
        res["esp_intake_bar"] = round(Pwf - rho*GRAVITY*(comm.z_ref - esp)/1e5, 1)
        res["dynamic_level_m"]= round(comm.z_ref - (Pwf - 1.0)*1e5/(rho*GRAVITY), 0)
        # --- per-layer flow split at the design rate -> FLOW-WEIGHTED MIXING T ---
        # The seven sands (48.7-52.9 C) commingle; the bottomhole flowing temperature
        # is the flow-weighted mean, NOT the deepest layer. The wellhead is then this
        # mix minus (small, at this high rate) Ramey wellbore conduction loss.
        Pwf_ref = comm.Pwf_at_q(target_flow_ls/1000.0)
        rates = comm.layer_rates_at_Pwf(Pwf_ref)                 # [(name, q_m3s)]
        Tmap  = {L.name: L.T_res_C for L in comm.layers}
        q_sum = sum(q for _, q in rates) or 1e-9
        T_mix = sum(q*Tmap[n] for n, q in rates)/q_sum
        res["mixing_T_C"] = round(T_mix, 2)
        res["layer_split"] = [dict(name=n, q_ls=round(q*1000,2), q_m3d=round(q*86400,0),
                                   share_pct=round(100*q/q_sum,1), T_C=Tmap[n])
                              for n, q in rates]
        # wellbore conduction loss (mix -> wellhead); column treated liquid-full (conservative)
        try:
            import wellbore_T as _W
            Twh = _W.wellhead_temperature(
                L_m=comm.z_ref, q_ls=target_flow_ls, T_bh_C=T_mix,
                geo_grad_K_m=NEWWELL_PRODUCER['thermal']['geo_gradient_K_m'],
                T_surf_C=NEWWELL_PRODUCER['thermal']['T_surface_C'],
                r_wb=0.125, rho_f=988.0, cp_f=4185.0,
                P_bh_bar=Pwf, P_wh_bar=NEWWELL_PRODUCER['operating']['WHP_bar'],
                GWR_m3m3=0.0, dyn_level_m=0.0, prod_time_yr=1.0)
            res["wellhead_T_C"] = round(float(Twh), 1)
        except Exception:
            res["wellhead_T_C"] = round(T_mix - 0.5, 1)         # tiny loss fallback
        res["source"] = "Ze-1 live import (main_ze1, new-well producer; mixing-T wellhead)"
    except Exception as e:
        res["import_note"] = f"(live import not active: {e}); using documented Ze-1 constants"
    return res


# ===========================================================================
# 3) ENGINEERING INTERFACE  -  ESP, injection pump, DHS, doublet
# ===========================================================================
SCEN_RETURN = {"A": 60.0, "B": 40.0, "C": 15.0}

def _silent(fn, *a, **k):
    with contextlib.redirect_stdout(io.StringIO()):
        return fn(*a, **k)

# --- surface-pipeline hydraulics (so flow changes are checked, not silently fixed) ---
def _mu_water(T_C):
    """Dynamic viscosity of water [Pa.s] (Vogel correlation, ~0-100 C)."""
    return 2.414e-5 * 10.0**(247.8 / ((T_C + 273.15) - 140.0))

def _pipe_dP(Q_m3s, D_m, L_m, rho, mu, eps_m=4.6e-5):
    """Darcy-Weisbach pressure drop with Colebrook friction for a round pipe."""
    if D_m <= 0 or Q_m3s <= 0:
        return dict(v_ms=0.0, Re=0.0, f=0.0, dP_bar=0.0, dP_Pa=0.0)
    A  = math.pi * D_m**2 / 4.0
    v  = Q_m3s / A
    Re = rho * v * D_m / max(mu, 1e-9)
    if Re < 2300.0:
        f = 64.0 / max(Re, 1e-9)
    else:
        f = 0.25 / (math.log10(eps_m/D_m/3.7 + 5.74/Re**0.9))**2
        for _ in range(50):
            fn = 1.0 / (-2.0*math.log10(eps_m/D_m/3.7 + 2.51/(Re*math.sqrt(f))))**2
            if abs(fn - f) < 1e-9:
                f = fn; break
            f = fn
    dP = f * (L_m / D_m) * 0.5 * rho * v**2
    return dict(v_ms=v, Re=Re, f=f, dP_bar=dP/1e5, dP_Pa=dP)

def _dn_for_velocity(Q_m3s, v_target=2.5):
    """Minimum inner diameter [m] to keep pipe velocity <= v_target."""
    if Q_m3s <= 0 or v_target <= 0:
        return 0.0
    return math.sqrt(4.0 * Q_m3s / (math.pi * v_target))

VEL_CAUTION_MS = 2.5    # comfort/erosion limit for liquid transmission lines
VEL_HIGH_MS    = 3.0    # hard caution (erosion / noise / excessive dP)
DPDL_CAUTION_PA_M = 150.0   # DH transmission pressure-gradient guideline (Pa per m)
DPDL_HIGH_PA_M    = 300.0   # clearly excessive gradient -> enlarge carrier pipe

def run_engineering(cfg, v3):
    # Zelina HP-assisted ambient loop: injection T = ambient return + cold approach (~13 C)
    reinj_T = cfg.dh_return_C + cfg.he_cold_approach_K
    import esp_geothermal_Ze1 as esp

    # ---- production ESP (lift = dynamic level; duty from V3) -----------------
    esp.PROD_FLOW_LS = v3["prod_flow_ls"]
    esp.PROD_TEMP_C  = v3["wellhead_T_C"]
    esp.PROD_DEPTH_M = v3["dynamic_level_m"]    # pumping lift height
    esp.PROD_BACKPRESSURE_BAR = 3.2
    prod = _silent(esp.calc_well_pump, "Prod", esp.PROD_DEPTH_M, esp.PROD_FLOW_LS,
                   esp.PROD_TEMP_C, esp.PROD_PUMP_EFF, esp.PROD_MOTOR_EFF,
                   esp.PROD_CABLE_EFF, esp.PROD_VSD_EFF, esp.PROD_PIPE_ID_M,
                   esp.PROD_PIPE_ROUGHNESS_M, esp.PROD_BACKPRESSURE_BAR, mode="production")

    # ---- injection pump (mirror well; required BHP from V3) ------------------
    esp.INJ_FLOW_LS = v3["prod_flow_ls"]
    esp.INJ_TEMP_C  = reinj_T
    esp.INJ_DEPTH_M = v3["inj_depth_m"]
    esp.INJ_RESERVOIR_BAR = v3["inj_bhp_bar"]
    esp.INJ_WELLHEAD_BAR  = 0.0
    inj = _silent(esp.calc_well_pump, "Inj", esp.INJ_DEPTH_M, esp.INJ_FLOW_LS,
                  esp.INJ_TEMP_C, esp.INJ_PUMP_EFF, esp.INJ_MOTOR_EFF,
                  esp.INJ_CABLE_EFF, esp.INJ_VSD_EFF, esp.INJ_PIPE_ID_M,
                  esp.INJ_PIPE_ROUGHNESS_M, esp.INJ_WELLHEAD_BAR, mode="injection",
                  reservoir_pressure_bar=esp.INJ_RESERVOIR_BAR)

    # ---- DHS: PHE + buried pipeline loss + delivered heat + circulator -------
    import geothermal_HE_Ze1 as dhs
    dhs.T_brine_in = v3["wellhead_T_C"]
    dhs.Q_brine    = v3["prod_flow_ls"]
    dhs.L_pipe     = cfg.city_distance_m
    dhs.D_inner    = cfg.dh_pipe_DN_m
    R = _silent(dhs.run_with_fallback, cfg.dh_scenario, SCEN_RETURN[cfg.dh_scenario],
                None, None, verbose=False) or {}
    # extract delivered heat / PHE area / circulator electrical power
    Q_HE = R.get("Q_HE", R.get("Q_W", None))
    if Q_HE is None:                      # reconstruct from brine cooling if absent
        rho = 970.0; cp = 4185.0
        Q_HE = v3["prod_flow_ls"]/1000.0 * rho * cp * (v3["wellhead_T_C"] - R.get("T_brine_out", reinj_T+4))
    loss_W = (R.get("Q_loss_sup", 0.0) or 0.0) + (R.get("Q_loss_ret", 0.0) or 0.0)
    delivered_W = max(Q_HE - loss_W, 0.0)
    phe = R.get("PHE", {})
    # Use the ACTUAL sized exchanger area (full plate selection, real U & LMTD) — this is
    # what gets installed and what the EUR/m2 CAPEX is based on. Fall back to the thermo-
    # dynamic required area, then a rough Q/(U*LMTD) estimate only if both are absent.
    phe_area = phe.get("A_total_m2") or phe.get("A_req_m2") or phe.get("area_m2") or phe.get("A_m2")
    if phe_area is None:
        U = 3500.0; LMTD = max(R.get("approach_hot", 5.0), 4.0)
        phe_area = Q_HE / (U * LMTD)
    # circulator electrical power from DHS pump duty
    dp = R.get("dp_pump_Pa", 0.0) or 0.0
    Qv = R.get("Q_vol", dhs.Q_brine/1000.0) or (dhs.Q_brine/1000.0)
    circ_hyd_W = dp * Qv
    circ_kW = circ_hyd_W / (0.75*0.95*0.97) / 1000.0
    brine_out = R.get("T_brine_out", reinj_T + 4.0)

    # --- Zelina override: brine cools from wellhead T to injection T (ambient loop) ---
    # The HE module heats the network toward brine T (its A/B/C convention); Zelina runs an
    # ambient HP-source loop (dh_supply/return = 15/10 C), so we set the geothermal (HP
    # evaporator) duty explicitly from the brine cooling and the computed injection T.
    brine_out = reinj_T                                   # injection T ~ dh_return + cold approach
    _rho_g, _cp_g = 988.0, 4185.0
    Q_HE = (v3["prod_flow_ls"]/1000.0) * _rho_g * _cp_g * (v3["wellhead_T_C"] - brine_out)
    delivered_W = max(Q_HE - loss_W, 0.0)                 # HP evaporator (geothermal) heat after pipe loss

    # ---- surface pipeline hydraulics (flow-driven velocity & pressure-drop check) ----
    # Brine reinjection line (cooled brine -> injection well): NOT modelled elsewhere,
    # so compute it here from the live flow + the user's DN. The DH carrier pipe
    # velocity / pipe pressure-drop are already produced by geothermal_HE (hyd_sup/ret);
    # we surface them and flag both lines if velocity exceeds safe limits.
    Q_brine_m3s = v3["prod_flow_ls"] / 1000.0
    rho_b = 985.0                                   # cooled brine ~ injection temperature
    mu_b  = 1.10 * _mu_water(brine_out)             # brine ~10% above pure water
    brine_hyd    = _pipe_dP(Q_brine_m3s, cfg.brine_pipe_DN_m, cfg.injection_distance_m, rho_b, mu_b)
    brine_DN_min = _dn_for_velocity(Q_brine_m3s, VEL_CAUTION_MS)

    hyd_sup = R.get("hyd_sup", {}) or {}
    hyd_ret = R.get("hyd_ret", {}) or {}
    dh_v    = hyd_sup.get("v", 0.0) or 0.0
    dh_dP_pipe_bar = ((hyd_sup.get("dp_total", 0.0) or 0.0) + (hyd_ret.get("dp_total", 0.0) or 0.0)) / 1e5
    Q_dh_m3s  = R.get("Q_vol", 0.0) or 0.0
    dh_DN_min = _dn_for_velocity(Q_dh_m3s, VEL_CAUTION_MS)

    warns = []
    def _flag(name, v, dn_min, D):
        if v >= VEL_HIGH_MS:
            warns.append(f"{name}: velocity {v:.2f} m/s exceeds {VEL_HIGH_MS:.1f} m/s "
                         f"(erosion / noise / high dP) - enlarge to >= DN{1000*dn_min:.0f} (now DN{1000*D:.0f}).")
        elif v >= VEL_CAUTION_MS:
            warns.append(f"{name}: velocity {v:.2f} m/s above the {VEL_CAUTION_MS:.1f} m/s comfort limit "
                         f"- consider >= DN{1000*dn_min:.0f} (now DN{1000*D:.0f}).")
    _flag("Brine reinjection line", brine_hyd["v_ms"], brine_DN_min, cfg.brine_pipe_DN_m)
    _flag("DH carrier pipe",        dh_v,             dh_DN_min,    cfg.dh_pipe_DN_m)
    def _flag_grad(name, dP_bar, L_total_m):
        if L_total_m <= 0:
            return
        g = dP_bar * 1e5 / L_total_m
        if g >= DPDL_HIGH_PA_M:
            warns.append(f"{name}: pressure gradient {g:.0f} Pa/m exceeds {DPDL_HIGH_PA_M:.0f} Pa/m "
                         f"(circulator head {dP_bar:.1f} bar over {L_total_m:.0f} m is high) - enlarge the carrier pipe.")
        elif g >= DPDL_CAUTION_PA_M:
            warns.append(f"{name}: pressure gradient {g:.0f} Pa/m above the {DPDL_CAUTION_PA_M:.0f} Pa/m guideline "
                         f"({dP_bar:.1f} bar over {L_total_m:.0f} m).")
    _flag_grad("DH carrier pipe", dh_dP_pipe_bar, 2.0 * cfg.city_distance_m)

    hydraulics = dict(
        brine_line=dict(D_m=cfg.brine_pipe_DN_m, L_m=cfg.injection_distance_m,
                        v_ms=brine_hyd["v_ms"], Re=brine_hyd["Re"],
                        dP_bar=brine_hyd["dP_bar"], DN_min_m=brine_DN_min),
        dh_line=dict(D_m=cfg.dh_pipe_DN_m, L_one_way_m=cfg.city_distance_m,
                     v_ms=dh_v, dP_pipe_bar=dh_dP_pipe_bar,
                     dP_circuit_bar=(dp/1e5), DN_min_m=dh_DN_min),
        warnings=warns)

    return dict(reinj_T=reinj_T,
                esp_kW=prod["surface_kW"], esp_TDH=prod["TDH_m"],
                inj_kW=(inj["surface_kW"] if inj.get("pump_needed", True) else 0.0),
                inj_TDH=inj.get("TDH_m", 0.0),
                Q_HE_kW=Q_HE/1000.0, delivered_kW=delivered_W/1000.0,
                pipe_loss_kW=loss_W/1000.0, phe_area_m2=phe_area,
                circ_kW=circ_kW, brine_out_C=brine_out,
                brine_line_v_ms=brine_hyd["v_ms"], brine_line_dP_bar=brine_hyd["dP_bar"],
                dh_line_v_ms=dh_v, dh_line_dP_pipe_bar=dh_dP_pipe_bar,
                hydraulics=hydraulics,
                _prod=prod, _inj=inj, _dhs=R)


# ===========================================================================
# 4) ENERGY (year 0) + doublet decline over life
# ===========================================================================
def energy_profile(cfg, v3, eng):
    import doublet_decline_Ze1 as dd
    rho, cp = 988.0, 4.185                       # kg/m3, kJ/kgK (harmonised with HE override)
    dT0 = v3["wellhead_T_C"] - eng["brine_out_C"]
    installed_kWth = v3["prod_flow_ls"]/1000.0 * rho * cp * dT0   # brine-side gross
    # consumer-side heat-pump upgrade: Q_consumer = Q_geo * SPF/(SPF-1); W_HP = Q_geo/(SPF-1)
    _spf = float(getattr(cfg, "hp_spf", 4.7))
    delivered_geo_kWth = eng["delivered_kW"]                       # HP evaporator (geothermal) heat
    hp_elec_kW     = delivered_geo_kWth/(_spf-1.0)                # consumer heat-pump electricity
    # Revenue basis follows who OPERATES the HPs (= who pays their electricity):
    #   plant operates (hp_elec_in_opex=True)  -> plant sells the UPGRADED useful heat
    #   consumer operates (hp_elec_in_opex=False) -> plant sells only the LOW-T geothermal heat
    if getattr(cfg, "hp_elec_in_opex", True):
        delivered_kWth = delivered_geo_kWth * _spf/(_spf-1.0)     # upgraded (HP output)
    else:
        delivered_kWth = delivered_geo_kWth                      # low-T geothermal heat sold as-is
    # doublet decline
    t_yr, T_prod = dd.doublet_temperature_decline(
        cfg.doublet_avg_flow_ls, cfg.doublet_spacing_m, eng["reinj_T"],
        v3["layers"], years=cfg.project_life_yr,
        rho_w=v3.get("jelic",{}).get("brine",(1000.,4184.))[0],
        cp_w=v3.get("jelic",{}).get("brine",(1000.,4184.))[1],
        rock_rho=v3.get("jelic",{}).get("rock",(2589.9,931.8,2.5))[0],
        rock_cp=v3.get("jelic",{}).get("rock",(2589.9,931.8,2.5))[1],
        imp_K=v3.get("jelic",{}).get("rock",(2589.9,931.8,2.5))[2])
    dT_t = np.maximum(T_prod - eng["reinj_T"], 1e-6)
    decline = dT_t / dT_t[0]                      # fraction vs year 0
    # annual delivered MWh per project year (1..life)
    yrs = np.arange(1, cfg.project_life_yr + 1)
    decl_yr = np.interp(yrs, t_yr, decline)
    delivered_MWh = delivered_kWth * cfg.FLH / 1000.0 * decl_yr
    # --- optional Barends (2010) comparison view (economics stay on GS above) ---
    barends = None
    if getattr(cfg, "doublet_method", "GS") == "GS+Barends":
        _j = v3.get("jelic", {})
        barends = dd.decline_layered(
            cfg.doublet_avg_flow_ls, cfg.doublet_spacing_m, eng["reinj_T"],
            v3["layers"], years=cfg.project_life_yr,
            rho_w=_j.get("brine",(1000.,4184.))[0], cp_w=_j.get("brine",(1000.,4184.))[1],
            rock_rho=_j.get("rock",(2589.9,931.8,2.5))[0], rock_cp=_j.get("rock",(2589.9,931.8,2.5))[1],
            imp_K=_j.get("rock",(2589.9,931.8,2.5))[2],
            a_L=getattr(cfg, "barends_dispersivity_m", 5.0),
            include_underburden=getattr(cfg, "barends_underburden", True))
    return dict(installed_kWth=installed_kWth, delivered_kWth=delivered_kWth,
                delivered_geo_kWth=delivered_geo_kWth, hp_elec_kW=hp_elec_kW,
                delivered_MWh_y0=delivered_kWth*cfg.FLH/1000.0,
                delivered_MWh=delivered_MWh, decline=decl_yr,
                T_prod_end=float(T_prod[-1]), t_yr=t_yr, T_prod=T_prod, barends=barends)


# ===========================================================================
# 5) CAPEX
# ===========================================================================
def build_capex(cfg, v3, eng):
    total_kVA = (eng["esp_kW"] + eng["inj_kW"] + eng["circ_kW"]) / 0.9 * 1.25  # +margin
    items = {}
    # PRODUCER = new well (drilled); INJECTOR = existing Ze-1 (only conversion cost)
    _prod_cost = cfg.prod_well_cost_eur if cfg.prod_well_cost_eur > 0 \
                 else cfg.well_cost_eur_per_m * cfg.prod_well_depth_m
    items["Production well (new, turnkey)"] = _prod_cost
    items["Injection well (Ze-1 conversion)"] = cfg.inj_well_cost_eur
    # consumer/central heat pumps sized on the delivered (upgraded) duty
    _hp_delivered_kW = eng["delivered_kW"] * cfg.hp_spf/(cfg.hp_spf-1.0)   # consumer (upgraded) duty
    items["Heat pumps (SPF %.1f)" % cfg.hp_spf] = (_hp_delivered_kW * cfg.hp_eur_per_kW
                                                   if cfg.hp_capex_in_scope else 0.0)
    items["ESP system (pump+cable+VSD+install)"] = (eng["esp_kW"]*cfg.esp_eur_per_kW
                                                    + cfg.esp_cable_eur_per_m*v3["esp_depth_m"]
                                                    + cfg.esp_install_eur)
    items["Injection pump system"]       = eng["inj_kW"]*cfg.injpump_eur_per_kW + cfg.injpump_install_eur
    items["PHE (single, sized for duty)"] = eng["phe_area_m2"] * cfg.phe_eur_per_m2
    items["DH circulator main+reserve"]  = 2 * max(eng["circ_kW"],1.0) * cfg.circ_eur_per_kW
    items["Energy plant/container @ prod"]= cfg.plant_prod_eur
    items["Container @ injection well"]  = cfg.plant_inj_eur
    items["DH pipeline (supply+return)"] = cfg.dh_pipe_eur_per_m * 2.0 * cfg.city_distance_m
    items["Brine reinjection pipeline"]  = cfg.brine_pipe_eur_per_m * cfg.injection_distance_m
    items["Grid connection (transformer+line+fee)"] = (total_kVA*cfg.transformer_eur_per_kVA
                                                       + cfg.mv_line_eur_per_m*cfg.grid_distance_m
                                                       + cfg.grid_connection_fee_eur)
    subtotal = sum(items.values())
    items["Engineering & design"]        = cfg.eng_pct * subtotal
    items["Contingency"]                 = cfg.contingency_pct * subtotal
    items["_TOTAL_"]                     = sum(items.values())
    items["_total_kVA_"]                 = total_kVA
    return items


# ===========================================================================
# 6) OPEX (year-1 basis) + scheduled replacements
# ===========================================================================
def build_opex(cfg, v3, eng, ener, capex):
    elec_MWh = (eng["esp_kW"] + eng["inj_kW"] + eng["circ_kW"]) * cfg.FLH / 1000.0
    surface_capex = (capex["ESP system (pump+cable+VSD+install)"]
                     + capex["Injection pump system"] + capex["PHE (single, sized for duty)"]
                     + capex["DH circulator main+reserve"]
                     + capex["Energy plant/container @ prod"] + capex["Container @ injection well"]
                     + capex["Grid connection (transformer+line+fee)"])
    well_capex = capex["Production well (new, turnkey)"] + capex["Injection well (Ze-1 conversion)"]
    op = {}
    hp_elec_MWh = (ener.get("hp_elec_kW", 0.0) * cfg.FLH / 1000.0) if cfg.hp_elec_in_opex else 0.0
    op["Electricity (ESP+inj+circ)"] = elec_MWh * cfg.elec_price_eur_MWhe
    op["Heat-pump electricity"]      = hp_elec_MWh * cfg.elec_price_eur_MWhe
    op["Grid demand charge"]         = capex["_total_kVA_"]*0.9 * cfg.grid_capacity_charge_eur_kW_yr
    op["Personnel (1 FTE)"]          = cfg.personnel_eur_yr
    op["Service & maintenance"]      = cfg.sm_pct_surface*surface_capex + cfg.sm_pct_wells*well_capex
    op["Chemicals / inhibitor / cleaning"] = cfg.chemicals_eur_yr
    op["Insurance"]                  = cfg.insurance_pct * capex["_TOTAL_"]
    op["Misc (telemetry/land/admin)"]= cfg.misc_opex_eur_yr
    op["_elec_MWh_"] = elec_MWh
    return op


# ===========================================================================
# 7) CASHFLOW + 8) METRICS
# ===========================================================================
def npv(rate, cfs):
    return sum(cf/(1.0+rate)**t for t, cf in enumerate(cfs))

def irr(cfs):
    from scipy.optimize import brentq
    try:
        return brentq(lambda r: npv(r, cfs), -0.95, 5.0, xtol=1e-6)
    except Exception:
        return float("nan")

def run_cashflow(cfg, v3, eng, ener, capex, opex):
    N = cfg.project_life_yr
    CAPEX = capex["_TOTAL_"]
    # depreciation: wells/pipeline 30 yr, equipment 15 yr (straight line)
    dep_long = (capex["Production well (new, turnkey)"]+capex["Injection well (Ze-1 conversion)"]
                + capex["DH pipeline (supply+return)"]+capex["Brine reinjection pipeline"]) / 30.0
    dep_short = (CAPEX - 30*dep_long if False else
                 (CAPEX - (capex["Production well (new, turnkey)"]+capex["Injection well (Ze-1 conversion)"]
                           +capex["DH pipeline (supply+return)"]+capex["Brine reinjection pipeline"]))) / 15.0
    # debt
    debt = cfg.debt_ratio * CAPEX
    equity = CAPEX - debt
    # annuity loan payment
    if debt > 0 and cfg.loan_interest > 0:
        i, n = cfg.loan_interest, cfg.loan_tenor_yr
        ann = debt * i / (1 - (1+i)**(-n))
    else:
        ann = debt / cfg.loan_tenor_yr if debt > 0 else 0.0

    rows = []
    bal = debt
    cum_disc = 0.0; payback_yr = None
    fcf = [-CAPEX]              # year 0 (project free cash flow, unlevered)
    equity_cf = [-equity]      # year 0 (levered, equity)
    co2_t = ener["delivered_MWh"]/cfg.gas_boiler_eff * cfg.ng_emission_t_MWh
    for y in range(1, N+1):
        esc_h = (1+cfg.heat_price_escal)**(y-1)
        esc_e = (1+cfg.elec_price_escal)**(y-1)
        esc_o = (1+cfg.opex_escal)**(y-1)
        delivered = ener["delivered_MWh"][y-1]
        # revenue
        rev_energy   = delivered * cfg.heat_price_eur_MWhth * esc_h
        rev_capacity = ener["installed_kWth"] * cfg.capacity_price_eur_kW_yr   # fixed contracted
        rev_co2      = (co2_t[y-1]*cfg.co2_price_eur_t) if cfg.co2_as_revenue else 0.0
        revenue = rev_energy + rev_capacity + rev_co2
        # opex (+ concession variable on delivered heat)
        concession = (cfg.field_area_km2*cfg.concession_fixed_eur_km2
                      + cfg.concession_var_pct*rev_energy)
        opex_y = (sum(v for k,v in opex.items() if not k.startswith("_"))*esc_e_if(k_is_elec=False)*esc_o
                  if False else
                  opex["Electricity (ESP+inj+circ)"]*esc_e
                  + (sum(v for k,v in opex.items() if not k.startswith("_") and k!="Electricity (ESP+inj+circ)"))*esc_o
                  + concession)
        # scheduled replacements (capex events)
        repl = 0.0
        if y % cfg.esp_replace_interval_yr == 0 and y < N:
            repl += eng["esp_kW"]*cfg.esp_eur_per_kW + cfg.esp_cable_eur_per_m*v3["esp_depth_m"]
        if y % cfg.injpump_replace_interval_yr == 0 and y < N:
            repl += eng["inj_kW"]*cfg.injpump_eur_per_kW
        # debt service
        interest = bal*cfg.loan_interest if y <= cfg.loan_tenor_yr and bal>0 else 0.0
        principal = (ann - interest) if y <= cfg.loan_tenor_yr and bal>0 else 0.0
        principal = min(principal, bal)
        bal = max(bal - principal, 0.0)
        # depreciation
        dep = dep_long + (dep_short if y <= 15 else 0.0)
        ebitda = revenue - opex_y
        ebt = ebitda - dep - interest
        tax = max(ebt, 0.0)*cfg.tax_rate
        net_income = ebt - tax
        # cash flows
        proj_fcf = ebitda - tax - repl                          # unlevered (for project NPV/IRR @ discount)
        eq_cf    = ebitda - tax - interest - principal - repl   # levered equity cash flow
        dscr = ebitda/(interest+principal) if (interest+principal) > 0 else float("inf")
        fcf.append(proj_fcf); equity_cf.append(eq_cf)
        # discounted payback (project)
        disc = proj_fcf/(1+cfg.discount_rate)**y
        cum_disc += disc
        if payback_yr is None and (cum_disc - CAPEX) >= 0:
            payback_yr = y
        rows.append(dict(year=y, delivered_MWh=delivered, revenue=revenue,
                         rev_energy=rev_energy, rev_capacity=rev_capacity, rev_co2=rev_co2,
                         opex=opex_y, concession=concession, repl=repl,
                         interest=interest, principal=principal, debt_bal=bal,
                         dep=dep, tax=tax, ebitda=ebitda, proj_fcf=proj_fcf,
                         equity_cf=eq_cf, dscr=dscr))
    # metrics
    proj_npv = npv(cfg.discount_rate, fcf)
    proj_irr = irr(fcf)
    eq_npv   = npv(cfg.discount_rate, equity_cf)
    eq_irr   = irr(equity_cf) if equity > 0 else float("nan")
    # simple payback
    cum = -CAPEX; simple_pb = None
    for y in range(1, N+1):
        cum += fcf[y]
        if simple_pb is None and cum >= 0: simple_pb = y
    # LCOH = (disc capex + disc opex) / disc delivered heat
    disc_costs = CAPEX + sum((rows[y-1]["opex"]+rows[y-1]["repl"])/(1+cfg.discount_rate)**y for y in range(1,N+1))
    disc_heat  = sum(rows[y-1]["delivered_MWh"]/(1+cfg.discount_rate)**y for y in range(1,N+1))
    lcoh = disc_costs/disc_heat if disc_heat>0 else float("nan")
    dscr_vals = [r["dscr"] for r in rows if math.isfinite(r["dscr"])]
    return dict(rows=rows, fcf=fcf, equity_cf=equity_cf,
                proj_npv=proj_npv, proj_irr=proj_irr, eq_npv=eq_npv, eq_irr=eq_irr,
                simple_payback=simple_pb, disc_payback=payback_yr, lcoh=lcoh,
                dscr_min=min(dscr_vals) if dscr_vals else float("inf"),
                dscr_avg=float(np.mean(dscr_vals)) if dscr_vals else float("inf"),
                debt=debt, equity=equity, ann=ann)

def esc_e_if(k_is_elec):   # tiny helper kept for clarity
    return 1.0


# ===========================================================================
# 9) RUN + REPORT
# ===========================================================================
def run(cfg=None):
    cfg = cfg or Config()
    # ---- DERIVE peak/design flow from the annual-average flow + operating duty ----
    # The doublet circulates at PEAK flow for `operating_months_per_yr` and is off the
    # rest of the year, so   avg = peak * (months/12)   ->   peak = avg / (months/12).
    # The PEAK flow sizes the reservoir IPR (Pwf, drawdown, injection BHP, layer split),
    # the ESP (depth/kW), the injection pump and the PHE. The AVERAGE flow is used only
    # by the Gringarten-Sauty thermal decline (see energy_profile()), which needs the
    # true time-averaged throughput to place the cold front correctly.
    _duty = max(cfg.operating_months_per_yr / 12.0, 1e-6)
    peak_flow_ls = cfg.doublet_avg_flow_ls / _duty
    v3  = get_v3_results(target_flow_ls=peak_flow_ls, reinj_T_C=SCEN_RETURN[cfg.dh_scenario], cfg=cfg)
    v3['avg_flow_ls']             = cfg.doublet_avg_flow_ls
    v3['operating_months_per_yr'] = cfg.operating_months_per_yr
    v3['duty_fraction']           = _duty
    v3['peak_flow_ls']            = peak_flow_ls   # == prod_flow_ls (design point)
    import wellbore_T, geothermo_props as gtp
    _rk = gtp.rock_props(v3['inj_depth_m'], SUBBASIN)
    _br = gtp.brine_props(T_BH, SALINITY)
    # Bottomhole flowing T = FLOW-WEIGHTED MIXING T of the 7 sands (computed in
    # get_v3_results from the per-layer split), NOT the deepest layer's static T.
    _T_bh_mix = v3.get('mixing_T_C', T_BH)
    v3['T_bh_C'] = _T_bh_mix
    v3['wellhead_T_C'] = round(wellbore_T.wellhead_temperature(
        L_m=v3['inj_depth_m'], q_ls=v3['prod_flow_ls'], T_bh_C=_T_bh_mix,
        geo_grad_K_m=GEO_GRAD, k_e=_rk[2], GWR_m3m3=GWR_M3M3,
        dyn_level_m=0.0,                     # column liquid-full (pumped); conservative
        P_bh_bar=v3['Pwf_bar'], P_wh_bar=3.0), 1)
    v3['jelic'] = dict(rock=_rk, brine=_br)
    eng = run_engineering(cfg, v3)
    ener= energy_profile(cfg, v3, eng)
    cap = build_capex(cfg, v3, eng)
    op  = build_opex(cfg, v3, eng, ener, cap)
    cf  = run_cashflow(cfg, v3, eng, ener, cap, op)
    return dict(cfg=cfg, v3=v3, eng=eng, ener=ener, capex=cap, opex=op, cf=cf)


def print_report(R):
    cfg, v3, eng, ener, cap, op, cf = (R["cfg"], R["v3"], R["eng"], R["ener"],
                                       R["capex"], R["opex"], R["cf"])
    L = "="*70
    print(L); print(" OsGT-1 (OSIJEK) GEOTHERMAL DOUBLET DH  -  ECONOMIC FEASIBILITY"); print(L)
    print(f" V3 source: {v3['source']}")
    print(f" Production: {v3['prod_flow_ls']:.1f} L/s, wellhead {v3['wellhead_T_C']:.0f} C, "
          f"drawdown {v3['drawdown_bar']:.0f} bar, ESP @ {v3['esp_depth_m']:.0f} m (intake {v3['esp_intake_bar']:.0f} bar)")
    print(f" Injection (mirror): BHP {v3['inj_bhp_bar']:.0f} bar -> pump {eng['inj_kW']:.0f} kW")
    print(f" DH scenario {cfg.dh_scenario} (return {SCEN_RETURN[cfg.dh_scenario]:.0f} C); brine cooled to {eng['brine_out_C']:.1f} C")
    print("-"*70)
    print(f" Installed thermal capacity : {ener['installed_kWth']/1000:8.3f} MWth (brine side)")
    print(f" Delivered heat (post-loss) : {ener['delivered_kWth']/1000:8.3f} MWth ; {ener['delivered_MWh_y0']:8.0f} MWh/yr (yr1)")
    print(f" Pipeline heat loss         : {eng['pipe_loss_kW']:8.1f} kW ; PHE area {eng['phe_area_m2']:.0f} m2")
    print(f" Pumps: ESP {eng['esp_kW']:.0f} kW, injection {eng['inj_kW']:.0f} kW, circulator {eng['circ_kW']:.1f} kW ; elec {op['_elec_MWh_']:.0f} MWh/yr")
    print(f" Doublet: T_prod {v3['wellhead_T_C']:.0f} C -> {ener['T_prod_end']:.1f} C at yr {cfg.project_life_yr} "
          f"(spacing {cfg.doublet_spacing_m:.0f} m, avg {cfg.doublet_avg_flow_ls:.1f} L/s)")
    print("-"*70); print(" CAPEX breakdown [EUR]")
    for k, v in cap.items():
        if not k.startswith("_"): print(f"   {k:<42s} {v:>14,.0f}")
    print(f"   {'TOTAL CAPEX':<42s} {cap['_TOTAL_']:>14,.0f}")
    print("-"*70); print(" OPEX (year 1) [EUR/yr]")
    for k, v in op.items():
        if not k.startswith("_"): print(f"   {k:<42s} {v:>14,.0f}")
    print(f"   {'Concession (yr1)':<42s} {cf['rows'][0]['concession']:>14,.0f}")
    print("-"*70); print(" REVENUE (year 1) [EUR/yr]")
    r0 = cf["rows"][0]
    print(f"   {'Energy (delivered heat)':<42s} {r0['rev_energy']:>14,.0f}")
    print(f"   {'Capacity payment':<42s} {r0['rev_capacity']:>14,.0f}")
    print(f"   {'CO2 credit'+('' if cfg.co2_as_revenue else ' (informational)'):<42s} {r0['rev_co2']:>14,.0f}")
    print("-"*70); print(" FINANCING & METRICS")
    print(f"   Debt {cfg.debt_ratio*100:.0f}% = {cf['debt']:,.0f} EUR @ {cfg.loan_interest*100:.1f}% / {cfg.loan_tenor_yr} yr ; equity {cf['equity']:,.0f}")
    print(f"   Discount rate              : {cfg.discount_rate*100:.1f}%")
    print(f"   Project NPV                : {cf['proj_npv']:>14,.0f} EUR")
    print(f"   Project IRR                : {cf['proj_irr']*100:>13.1f} %")
    if cf['equity'] > 0:
        print(f"   Equity NPV / IRR           : {cf['eq_npv']:,.0f} EUR / {cf['eq_irr']*100:.1f} %")
    print(f"   Simple / discounted payback: {cf['simple_payback']} / {cf['disc_payback']} yr")
    print(f"   LCOH                       : {cf['lcoh']:>13.2f} EUR/MWhth   (vs price {cfg.heat_price_eur_MWhth:.0f})")
    print(f"   DSCR min / avg             : {cf['dscr_min']:.2f} / {cf['dscr_avg']:.2f}")
    print(L)


def tornado(R, save_dir):
    """One-at-a-time +/-20% sensitivity of project NPV."""
    base = R["cf"]["proj_npv"]
    levers = {
        "Heat price":        ("heat_price_eur_MWhth", 0.20),
        "Electricity price": ("elec_price_eur_MWhe", 0.20),
        "FLH":               ("FLH", 0.20),
        "CAPEX (well cost/m)":("well_cost_eur_per_m", 0.20),
        "Discount rate":     ("discount_rate", 0.20),
        "Doublet spacing":   ("doublet_spacing_m", 0.20),
        "Capacity price":    ("capacity_price_eur_kW_yr", 0.20),
    }
    res = []
    for name, (attr, pct) in levers.items():
        lo = Config(**{**asdict(R["cfg"]), attr: getattr(R["cfg"], attr)*(1-pct)})
        hi = Config(**{**asdict(R["cfg"]), attr: getattr(R["cfg"], attr)*(1+pct)})
        try:
            nlo = run(lo)["cf"]["proj_npv"]; nhi = run(hi)["cf"]["proj_npv"]
        except Exception:
            nlo = nhi = base
        res.append((name, nlo, nhi))
    try:
        import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
        res.sort(key=lambda x: abs(x[2]-x[1]))
        names = [r[0] for r in res]; los = [r[1] for r in res]; his = [r[2] for r in res]
        fig, ax = plt.subplots(figsize=(9,5), dpi=140)
        y = np.arange(len(names))
        for i,(n,lo,hi) in enumerate(res):
            ax.barh(i, hi-lo, left=min(lo,hi), color="#2E75B6", alpha=0.8)
        ax.axvline(base, color="k", ls="--", lw=1, label=f"base NPV {base/1e6:.2f} M€")
        ax.set_yticks(y); ax.set_yticklabels(names); ax.set_xlabel("Project NPV [EUR]")
        ax.set_title("Tornado - project NPV sensitivity (+/-20%)"); ax.legend()
        fig.tight_layout(); p=os.path.join(save_dir,"economy_tornado.png"); fig.savefig(p); plt.close(fig)
        return p
    except Exception as e:
        return f"(tornado plot skipped: {e})"


def plots(R, save_dir):
    import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
    cf = R["cf"]; ener = R["ener"]; cfg = R["cfg"]
    out = []
    # cumulative discounted cash flow (payback)
    fig, ax = plt.subplots(figsize=(8,4.5), dpi=140)
    cum = np.cumsum([cf["fcf"][0]] + [cf["fcf"][y]/(1+cfg.discount_rate)**y for y in range(1,cfg.project_life_yr+1)])
    ax.plot(range(0,cfg.project_life_yr+1), np.array(cum)/1e6, marker="o", ms=3, color="#1F3864")
    ax.axhline(0, color="k", lw=0.8); ax.set_xlabel("Year"); ax.set_ylabel("Cumulative discounted cash flow [M€]")
    ax.set_title("Discounted payback"); ax.grid(alpha=0.3)
    fig.tight_layout(); p=os.path.join(save_dir,"economy_payback.png"); fig.savefig(p); plt.close(fig); out.append(p)
    # temperature decline
    from matplotlib.ticker import FormatStrFormatter
    fig, ax = plt.subplots(figsize=(8,5.8), dpi=140)
    ax.plot(ener["t_yr"], ener["T_prod"], color="#C00000")
    ax.set_xlabel("Year"); ax.set_ylabel("Producer temperature [\u00b0C]")
    ax.set_title(f"Gringarten-Sauty doublet decline (spacing {cfg.doublet_spacing_m:.0f} m, {cfg.doublet_avg_flow_ls:.0f} L/s avg)")
    _Tp = np.asarray(ener["T_prod"]); _lo, _hi = float(_Tp.min()), float(_Tp.max())
    if _hi - _lo < 10.0:
        _mid = 0.5*(_lo+_hi); _lo, _hi = _mid-6.0, _mid+6.0
    ax.set_ylim(_lo-1.0, _hi+1.0)
    ax.ticklabel_format(style="plain", axis="y", useOffset=False)
    ax.yaxis.set_major_formatter(FormatStrFormatter("%.1f"))
    ax.grid(alpha=0.3); fig.tight_layout()
    p=os.path.join(save_dir,"economy_Tdecline.png"); fig.savefig(p); plt.close(fig); out.append(p)
    return out


def to_excel(R, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as e:
        return f"(openpyxl missing: {e})"
    cfg, cap, op, cf, ener, eng, v3 = (R["cfg"],R["capex"],R["opex"],R["cf"],R["ener"],R["eng"],R["v3"])
    wb = Workbook(); hd = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="1F3864")
    def sheet(title, rows):
        ws = wb.create_sheet(title)
        for r,row in enumerate(rows,1):
            for c,val in enumerate(row,1):
                cell = ws.cell(r,c,val)
                if r==1: cell.font=hd; cell.fill=fill
        return ws
    wb.remove(wb.active)
    sheet("Summary", [["Metric","Value","Unit"],
        ["Production flow", v3["prod_flow_ls"], "L/s"],
        ["Installed capacity", round(ener["installed_kWth"]/1000,3), "MWth"],
        ["Delivered heat yr1", round(ener["delivered_MWh_y0"]), "MWh/yr"],
        ["Total CAPEX", round(cap["_TOTAL_"]), "EUR"],
        ["Project NPV", round(cf["proj_npv"]), "EUR"],
        ["Project IRR", round(cf["proj_irr"]*100,1), "%"],
        ["Discounted payback", cf["disc_payback"], "yr"],
        ["LCOH", round(cf["lcoh"],2), "EUR/MWhth"],
        ["DSCR min", round(cf["dscr_min"],2), "-"]])
    sheet("CAPEX", [["Item","EUR"]] + [[k,round(v)] for k,v in cap.items() if not k.startswith("_")]
                   + [["TOTAL", round(cap["_TOTAL_"])]])
    sheet("OPEX_yr1", [["Item","EUR/yr"]] + [[k,round(v)] for k,v in op.items() if not k.startswith("_")])
    cfh = ["year","delivered_MWh","revenue","opex","concession","repl","interest","principal",
           "debt_bal","dep","tax","ebitda","proj_fcf","equity_cf","dscr"]
    sheet("Cashflow", [cfh] + [[round(r[k],1) if isinstance(r[k],float) else r[k] for k in cfh] for r in cf["rows"]])
    wb.save(path); return path


if __name__ == "__main__":
    R = run()
    print_report(R)
    save_dir = os.environ.get("ECON_OUT", THIS_DIR)
    xp = to_excel(R, os.path.join(save_dir, "OsGT1_economy.xlsx"))
    pp = plots(R, save_dir)
    tp = tornado(R, save_dir)
    try:
        import doublet_viz
        _mp = max(R['v3']['layers'], key=lambda L: L['k'])
        vz = doublet_viz.visualize_doublet(
            dict(h=_mp['h'], k=_mp['k'], name=_mp.get('name','reservoir')),
            R['cfg'].doublet_avg_flow_ls, R['cfg'].doublet_spacing_m,
            R['eng']['reinj_T'], _mp['T0'], phi_doublet=_mp['phi'],
            well_name=WELL_NAME, save_dir=save_dir)
        print(' Viz   :', vz)
    except Exception as e:
        print(' Viz skipped:', e)
    print(f"\n Excel : {xp}\n Plots : {pp}\n Tornado: {tp}")
