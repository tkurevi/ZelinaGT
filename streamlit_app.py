"""
streamlit_app.py  -  Web front-end for the Ze-1 (Zelina-1) geothermal techno-economic model.
Tabs: Reservoir & Well · Pumps · DHS · Economy & Finance.
Imports the existing model unchanged; surfaces its data as tables and figures.
Run locally:   streamlit run streamlit_app.py
"""
import os, io, math, tempfile, dataclasses
import matplotlib
matplotlib.use("Agg")
import matplotlib.figure
import numpy as np
import pandas as pd
import streamlit as st

import economy_Ze1 as M
import doublet_decline_Ze1 as DD
import doublet_viz
import app_figures as AF
import geothermal_HE_Ze1 as GHE
from main_ze1 import build_commingled, CALIBRATED_CONFIG as RES_CFG
import esp_geothermal_Ze1 as ESPMOD

st.set_page_config(page_title="Geothermal Techno-Economic Model — Ze-1 (Zelina-1)", layout="wide")
st.title("Geothermal Techno-Economic Model — Ze-1 (Zelina-1)")
st.caption("Doublet district heating · reservoir model + ESP + DHS + economics (live)")

DISPLAY_W = 520           # uniform on-screen width for every figure


# --------------------------- formatting helpers ----------------------------
def fmt_num(x):
    if isinstance(x, bool) or not isinstance(x, (int, float)):
        return x
    if x != x:
        return "—"
    a = abs(x)
    if a == 0:
        return "0"
    if a >= 100:
        return f"{x:,.1f}"
    if a >= 1:
        return f"{x:,.2f}"
    d = max(3, 1 - int(math.floor(math.log10(a))))
    return f"{x:.{d}f}"


def _fmt_frame(df):
    df = df.copy()
    try:
        return df.map(fmt_num)        # pandas >= 2.1 (Streamlit Cloud)
    except AttributeError:
        return df.applymap(fmt_num)   # older pandas


def show_df(df, **kw):
    st.dataframe(_fmt_frame(df), use_container_width=True, hide_index=True, **kw)


def clean(d):
    return {k: v for k, v in d.items() if not str(k).startswith("_")}


# ----------------------------- input help text -----------------------------
HELP = {
 "dh_scenario": "Ambient HP-source loop operating point. A = 15/10 °C supply/return (primary). B = 30/20 °C (optional warmer variant). Brine is cooled to (return + cold approach) for reinjection.",
 "hp_capex_in_scope": "Tick to include consumer/central heat-pump CAPEX (€/kW of delivered heat). Untick if the heat pumps are owned by the consumers.",
 "hp_elec_in_opex": "Tick if the PLANT operates the heat pumps: their electricity is an OPEX AND the plant sells the upgraded useful heat. Untick ⇒ consumers operate them and the plant sells only the low-T geothermal heat.",
 "hp_spf": "Seasonal performance factor of the consumer heat pumps. Q_delivered = Q_geo·SPF/(SPF−1); HP electricity = Q_geo/(SPF−1).",
 "hp_eur_per_kW": "Heat-pump CAPEX per kW of delivered (upgraded) thermal power. Used only if HP CAPEX is in scope.",
 "prod_well_depth_m": "Drilled depth of the NEW production well (m). 9 m below the S7 base (891 m).",
 "doublet_avg_flow_ls": "Annual-AVERAGE circulation flow (L/s). Drives the Gringarten–Sauty cold-front decline. The PEAK/design flow that sizes the reservoir IPR, ESP, pumps and PHE is DERIVED from this and the operating months: peak = avg ÷ (operating_months/12).",
 "operating_months_per_yr": "Months per year the doublet runs at peak flow (the rest = off). Sets the duty: peak/design flow = average ÷ (months/12). E.g. 26.25 L/s average over 9 months → 35.0 L/s peak. Changing this (or the average) re-sizes the whole reservoir + production engineering; the average alone still governs the thermal decline.",
 "doublet_spacing_m": "Distance between producer and injector (m). Larger spacing delays thermal breakthrough.",
 "doublet_method": "Thermal-decline model. GS = Gringarten–Sauty (drives the economics). GS+Barends additionally plots a Barends-2010 longitudinal-dispersion comparison figure (economics stay on GS).",
 "barends_dispersivity_m": "Barends longitudinal dispersivity a_L per layer (m). Typical within-layer 3–10 m. Only used when doublet_method = GS+Barends. Larger a_L smears the front (earlier breakthrough, gentler decline).",
 "barends_underburden": "Barends bleeding into cap+base rock (h_eff = 2x) vs cap only. Only used when doublet_method = GS+Barends.",
 "FLH": "Full-load hours per year (h/yr): equivalent hours at full output the plant sells heat.",
 "injectivity_multiplier": "Injectivity vs productivity ratio (–). 1.0 = injector as easy as producer; >1 = easier injection.",
 "heat_price_eur_MWhth": "Heat selling price (€ per MWh thermal).",
 "elec_price_eur_MWhe": "Electricity price for pumps (€ per MWh electric).",
 "capacity_price_eur_kW_yr": "District-heating capacity payment (€ per kW thermal per year) — counted as revenue.",
 "co2_price_eur_t": "CO₂ price (€ per tonne), used if the CO₂ credit is counted.",
 "co2_as_revenue": "Tick to count avoided-CO₂ credit as revenue.",
 "gas_boiler_eff": "Reference gas-boiler efficiency for the CO₂-avoided calculation. Fraction: 85 % = 0.85.",
 "ng_emission_t_MWh": "Natural-gas emission factor (tonnes CO₂ per MWh of gas).",
 "project_life_yr": "Economic project life (years).",
 "discount_rate": "Real discount rate for NPV. Enter as a fraction: 6 % = 0.06.",
 "debt_ratio": "Share of CAPEX financed by debt. 0 = all equity, 1 = all debt.",
 "loan_interest": "Loan interest rate. Fraction: 4.5 % = 0.045.",
 "loan_tenor_yr": "Loan repayment period (years).",
 "tax_rate": "Corporate tax rate. Fraction: 18 % = 0.18.",
 "inflation": "General inflation. Real analysis → 0. Fraction: 2 % = 0.02.",
 "heat_price_escal": "Annual real escalation of the heat price. Fraction: 1 % = 0.01.",
 "elec_price_escal": "Annual real escalation of electricity price. Fraction: 1 % = 0.01.",
 "opex_escal": "Annual real escalation of OPEX. Fraction: 1 % = 0.01.",
 "city_distance_m": "One-way distance well → city (m). Pipeline billed both ways (supply + return).",
 "injection_distance_m": "Distance well → injection well (m) for the single brine line.",
 "grid_distance_m": "Distance to the nearest medium-voltage grid connection point (m).",
 "dh_pipe_DN_m": "District-heating carrier pipe inner diameter (m). 0.2101 ≈ DN200.",
 "brine_pipe_DN_m": "Injection (brine) line inner diameter (m). 0.20 ≈ DN200.",
 "well_cost_eur_per_m": "Turnkey drilling & completion cost (€ per metre of well).",
 "prod_well_cost_eur": "Override production-well cost (€). 0 = use €/m × depth (or already paid).",
 "inj_well_depth_m": "Injection-well depth (m), for its turnkey cost.",
 "esp_eur_per_kW": "ESP pump+motor cost (€ per surface kW).",
 "esp_cable_eur_per_m": "Downhole ESP power-cable cost (€ per metre).",
 "esp_install_eur": "ESP installation service (lump sum €).",
 "injpump_eur_per_kW": "Injection-pump cost (€ per kW).",
 "injpump_install_eur": "Injection-pump installation (lump sum €).",
 "phe_eur_per_m2": "Plate heat-exchanger cost (€ per m² of plate area).",
 "circ_eur_per_kW": "District-heating circulation-pump cost (€ per kW).",
 "plant_prod_eur": "Energy-plant / container at the production well (€).",
 "plant_inj_eur": "Container at the injection well (€).",
 "dh_pipe_eur_per_m": "Buried pre-insulated DH pipeline cost (€ per metre of trench).",
 "brine_pipe_eur_per_m": "Buried brine reinjection line cost (€ per metre).",
 "transformer_eur_per_kVA": "Transformer cost (€ per kVA).",
 "mv_line_eur_per_m": "Medium-voltage electrical supply line cost (€ per metre), × grid distance.",
 "grid_connection_fee_eur": "One-off grid-connection fee (€).",
 "eng_pct": "Engineering & design, as a fraction of the rest of CAPEX. 8 % = 0.08.",
 "contingency_pct": "Contingency, as a fraction of CAPEX. 12 % = 0.12.",
 "personnel_eur_yr": "Annual personnel cost (€/yr).",
 "grid_capacity_charge_eur_kW_yr": "Grid demand (capacity) charge — a COST — € per connected kW per year.",
 "sm_pct_surface": "Annual service & maintenance of surface equipment, as a fraction of surface CAPEX. 3 % = 0.03. (This is OPEX.)",
 "sm_pct_wells": "Annual service & maintenance of the wells, as a fraction of well CAPEX. 1 % = 0.01. (This is OPEX.)",
 "chemicals_eur_yr": "Scaling/corrosion inhibitor + cleaning chemicals (€/yr).",
 "insurance_pct": "Annual insurance, as a fraction of total CAPEX. 0.5 % = 0.005.",
 "misc_opex_eur_yr": "Telemetry, land, admin and other miscellaneous OPEX (€/yr).",
 "esp_replace_interval_yr": "ESP replacement interval (years).",
 "injpump_replace_interval_yr": "Injection-pump replacement interval (years).",
 "field_area_km2": "Exploitation/concession field area (km²).",
 "concession_fixed_eur_km2": "Fixed concession fee (€ per km² per year).",
 "concession_var_pct": "Variable concession fee, as a fraction of energy revenue. 3 % = 0.03.",
}
FRACTIONS = {"discount_rate", "debt_ratio", "loan_interest", "tax_rate", "inflation",
             "heat_price_escal", "elec_price_escal", "opex_escal", "eng_pct", "contingency_pct",
             "sm_pct_surface", "sm_pct_wells", "insurance_pct", "concession_var_pct", "gas_boiler_eff"}

GROUPS = {
    "Operation": ["dh_scenario", "doublet_avg_flow_ls", "operating_months_per_yr", "doublet_spacing_m", "doublet_method", "barends_dispersivity_m", "barends_underburden", "FLH", "injectivity_multiplier"],
    "Energy prices & revenue": ["heat_price_eur_MWhth", "elec_price_eur_MWhe", "capacity_price_eur_kW_yr", "co2_price_eur_t", "co2_as_revenue", "hp_capex_in_scope", "hp_elec_in_opex", "hp_spf", "gas_boiler_eff", "ng_emission_t_MWh"],
    "Financing": ["project_life_yr", "discount_rate", "debt_ratio", "loan_interest", "loan_tenor_yr", "tax_rate", "inflation", "heat_price_escal", "elec_price_escal", "opex_escal"],
    "Distances & pipes": ["city_distance_m", "injection_distance_m", "grid_distance_m", "dh_pipe_DN_m", "brine_pipe_DN_m"],
    "CAPEX unit costs": ["well_cost_eur_per_m", "prod_well_cost_eur", "prod_well_depth_m", "inj_well_depth_m", "inj_well_cost_eur", "hp_eur_per_kW", "esp_eur_per_kW", "esp_cable_eur_per_m", "esp_install_eur", "injpump_eur_per_kW", "injpump_install_eur", "phe_eur_per_m2", "circ_eur_per_kW", "plant_prod_eur", "plant_inj_eur", "dh_pipe_eur_per_m", "brine_pipe_eur_per_m", "transformer_eur_per_kVA", "mv_line_eur_per_m", "grid_connection_fee_eur", "eng_pct", "contingency_pct"],
    "OPEX & replacement": ["personnel_eur_yr", "grid_capacity_charge_eur_kW_yr", "sm_pct_surface", "sm_pct_wells", "chemicals_eur_yr", "insurance_pct", "misc_opex_eur_yr", "esp_replace_interval_yr", "injpump_replace_interval_yr", "field_area_km2", "concession_fixed_eur_km2", "concession_var_pct"],
}
FIELDS = {f.name: f for f in dataclasses.fields(M.Config)}


def in_fmt(d):
    a = abs(d)
    if a >= 100: return "%.1f"
    if a >= 1 or a == 0: return "%.2f"
    return "%.4f"


st.sidebar.header("Inputs")
vals = {}
for group, names in GROUPS.items():
    with st.sidebar.expander(group, expanded=(group == "Operation")):
        for nm in names:
            f = FIELDS[nm]; d = f.default; h = HELP.get(nm, "")
            if nm == "dh_scenario":
                vals[nm] = st.selectbox(nm, ["A", "B"], help=h)
            elif nm == "doublet_method":
                vals[nm] = st.selectbox(nm, ["GS", "GS+Barends"], index=0, help=h)
            elif nm == "debt_ratio":
                vals[nm] = st.slider(nm, 0.0, 1.0, float(d), 0.05, help=h)
            elif isinstance(d, bool):
                vals[nm] = st.checkbox(nm, value=bool(d), help=h)
            elif isinstance(d, int):
                vals[nm] = st.number_input(nm, value=int(d), step=1, help=h)
            elif nm in FRACTIONS:
                vals[nm] = st.number_input(nm, value=float(d), min_value=0.0, max_value=1.0,
                                           step=0.005, format="%.4f", help=h)
            else:
                step = 1.0 if abs(d) < 1000 else 1000.0
                vals[nm] = st.number_input(nm, value=float(d), step=float(step), format=in_fmt(d), help=h)

run_clicked = st.sidebar.button("▶  Run model", type="primary")


def build_excel(R):
    from openpyxl import Workbook
    wb = Workbook(); cap, op, cf = R["capex"], R["opex"], R["cf"]
    ws = wb.active; ws.title = "CAPEX"; ws.append(["item", "EUR"])
    for k, v in clean(cap).items():
        if isinstance(v, (int, float)): ws.append([k, round(v)])
    ws2 = wb.create_sheet("OPEX"); ws2.append(["item", "EUR/yr"])
    for k, v in clean(op).items():
        if isinstance(v, (int, float)): ws2.append([k, round(v)])
    ws3 = wb.create_sheet("Cashflow"); rows = cf.get("rows", [])
    if rows:
        ws3.append(list(rows[0].keys()))
        for r in rows: ws3.append([round(x, 2) if isinstance(x, (int, float)) else x for x in r.values()])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


if not run_clicked:
    st.info("Set inputs in the sidebar and press **Run model**.")
    st.stop()

# ---------------------------------------------------------------------------
cfg = M.Config(**vals)
with st.spinner("Running reservoir + engineering + economics…"):
    R = M.run(cfg)
    cf, ener, eng, v3, cap, op = R["cf"], R["ener"], R["eng"], R["v3"], R["capex"], R["opex"]
    tmp = tempfile.mkdtemp()
capex_total = cap.get("_TOTAL_", sum(v for v in clean(cap).values() if isinstance(v, (int, float))))

tab_res, tab_pump, tab_dhs, tab_eco = st.tabs(
    ["🪨 Reservoir & Well", "⚙️ Pumps", "♨️ DHS", "💶 Economy & Finance"])

# ============================ RESERVOIR & WELL =============================
with tab_res:
    st.subheader("Producing layers")
    rows = []
    _v3L = v3.get("layers", [])
    for _i, L in enumerate(RES_CFG["layers"]):
        _phi = L.get("porosity")
        if _phi is None and _i < len(_v3L):
            _phi = _v3L[_i].get("phi")
        rows.append({
            "Layer": str(L.get("name", "")).capitalize(),
            "Roof depth (m)": L.get("top_depth_m"),
            "Bottom depth (m)": L.get("bottom_depth_m"),
            "Net thickness, h (m)": L.get("h_net_m"),
            "Permeability, k (mD)": L.get("k_md"),
            "Porosity, φ (–)": _phi,
            "Reservoir T (°C)": L.get("T_res_C"),
            "Reservoir P (bar)": L.get("P_res_bar"),
        })
    show_df(pd.DataFrame(rows))
    st.caption("Porosity shown is the reservoir/matrix value; for fractured or vuggy reservoirs the "
               "doublet thermal model may use a different effective porosity.")

    st.subheader("Reservoir & fluid properties")
    sal_gL = M.SALINITY / 1000.0
    props = {
        "Reference depth, z_ref (m)": v3.get("inj_depth_m"),
        "Bottomhole T, mixed (°C)": v3.get("T_bh_C", M.T_BH),
        "Wellhead T, Ramey (°C)": v3.get("wellhead_T_C"),
        "Static reservoir P (bar)": v3.get("static_bar"),
        "Flowing BHP, Pwf (bar)": v3.get("Pwf_bar"),
        "Drawdown (bar)": v3.get("drawdown_bar"),
        "Annual-average flow, input (L/s)": v3.get("avg_flow_ls", cfg.doublet_avg_flow_ls),
        "Operating months/yr": v3.get("operating_months_per_yr", cfg.operating_months_per_yr),
        "Peak / design flow, derived (L/s)": v3.get("peak_flow_ls", v3.get("prod_flow_ls")),
        "Geothermal gradient (K/m)": M.GEO_GRAD,
        "Sub-basin": M.SUBBASIN,
        "Gas-water ratio, GWR (–)": M.GWR_M3M3,
        "Salinity (g/L NaCl)": sal_gL,
    }
    show_df(pd.DataFrame(props.items(), columns=["property", "value"]))
    st.caption(
        f"Peak/design flow = average ÷ (operating months ÷ 12) = "
        f"{v3.get('avg_flow_ls', cfg.doublet_avg_flow_ls):.2f} ÷ "
        f"({v3.get('operating_months_per_yr', cfg.operating_months_per_yr)}/12) = "
        f"**{v3.get('peak_flow_ls', v3.get('prod_flow_ls')):.2f} L/s**. "
        "Flowing BHP, drawdown, injection BHP, ESP depth/kW, pumps and PHE are all sized at this peak flow; "
        "the annual-average flow drives only the Gringarten–Sauty thermal decline.")

    st.subheader("Well construction & geometry")
    bottom = max(L.get("bottom_depth_m", 0) for L in RES_CFG["layers"])
    well_rows = [
        {"Element": "Production tubing", "Roof depth (m)": 0.0, "Bottom depth (m)": v3.get("inj_depth_m"),
         "Inner Ø (mm)": RES_CFG["well"].get("tubing_ID", 0) * 1000,
         "Outer Ø (mm)": RES_CFG["well"].get("tubing_OD", 0) * 1000},
    ]
    for L in RES_CFG["layers"]:
        well_rows.append({"Element": f"Producing interval — {L.get('name','')}",
                          "Roof depth (m)": L.get("top_depth_m"), "Bottom depth (m)": L.get("bottom_depth_m"),
                          "Inner Ø (mm)": None, "Outer Ø (mm)": None})
    well_rows.append({"Element": "Total / reference datum", "Roof depth (m)": None,
                      "Bottom depth (m)": v3.get("inj_depth_m"), "Inner Ø (mm)": None, "Outer Ø (mm)": None})
    show_df(pd.DataFrame(well_rows))
    st.caption("Note: detailed casing-string program (surface/intermediate/production casing diameters & shoe "
               "depths) is not stored in the model config; add it here from the well completion report when available.")

    st.subheader("Doublet — thermal decline & cold front")
    figs = M.plots(R, tmp)                       # [payback, Tdecline(GS)]
    gs = next((f for f in figs if "Tdecline" in f), None)
    if gs and os.path.exists(gs):
        st.image(gs, width=DISPLAY_W, caption="Producer temperature decline (Gringarten–Sauty)")
    ba = ener.get("barends")
    if ba is not None:
        try:
            ba_png = DD.plot_barends_comparison(ba, os.path.join(tmp, "barends_compare.png"),
                                                well_label="Ze-1 (new producer)")
            st.image(ba_png, width=int(DISPLAY_W*1.5),
                     caption=("GS vs Barends (2010) — longitudinal-dispersion comparison  "
                              f"(a_L = {ba['a_L']:.1f} m, "
                              f"{'cap+base' if ba['include_underburden'] else 'cap-only'} bleeding)"))
            _b1, _b2, _b3 = st.columns(3)
            _b1.metric("Breakthrough (0.1%)", f"{ba['t_BT_BA']:.1f} yr",
                       f"{ba['t_BT_BA']-ba['t_BT_GS']:+.1f} yr vs G&S", delta_color="off")
            _b2.metric(f"{ba['ba_pct']:.0f}%-decline", f"{ba['t_BA_decline']:.1f} yr",
                       f"{ba['t_BA_decline']-ba['t_GS_decline']:+.1f} yr vs G&S", delta_color="off")
            _b3.metric("End-T (100 yr)", f"{ba['T_mix_ba'][-1]:.1f} °C",
                       f"{ba['T_mix_ba'][-1]-ba['T_mix_gs'][-1]:+.2f} °C vs G&S", delta_color="off")
            st.caption("Economics use the Gringarten–Sauty curve above; Barends is a comparison view only.")
        except Exception as _e:
            st.warning(f"Barends comparison skipped: {_e}")
    try:
        _mp = max(v3["layers"], key=lambda L: L["k"])
        viz = doublet_viz.visualize_doublet(dict(h=_mp["h"], k=_mp["k"], name=_mp.get("name", "reservoir")),
              cfg.doublet_avg_flow_ls, cfg.doublet_spacing_m, eng["reinj_T"], _mp["T0"],
              phi_doublet=_mp["phi"], well_name="Ze-1", save_dir=tmp)
        cold = [p for p in viz if "coldfront" in p]
        if cold:
            st.image(cold[0], width=DISPLAY_W, caption="Cold-front advance & swept area")
    except Exception as e:
        st.warning(f"Cold-front viz skipped: {e}")
    st.caption(f"Reservoir feed: {v3.get('source','n/a')}")

# ================================= PUMPS ===================================
with tab_pump:
    st.subheader("Pump sizing (duty)")
    st.caption(
        f"Sized at the peak/design flow **{v3.get('peak_flow_ls', v3.get('prod_flow_ls')):.2f} L/s** "
        f"(= {v3.get('avg_flow_ls', cfg.doublet_avg_flow_ls):.2f} L/s average ÷ "
        f"{v3.get('operating_months_per_yr', cfg.operating_months_per_yr)}/12 months). "
        "Higher flow → lower Pwf → deeper dynamic level → deeper ESP setting → higher TDH and kW.")
    prod, inj = eng.get("_prod", {}), eng.get("_inj", {})
    esp_tbl = {
        "ESP setting depth (m)": v3.get("esp_depth_m"),
        "Dynamic fluid level (m)": v3.get("dynamic_level_m"),
        "Wellhead T, Ramey (°C)": v3.get("wellhead_T_C"),
        "Intake NPSHa (m)": prod.get("NPSHa_m"),
        "Total dynamic head (m)": prod.get("TDH_m"),
        "Tubing velocity (m/s)": prod.get("vel_m_s"),
        "Reynolds (–)": prod.get("Re"),
        "Hydraulic power (kW)": prod.get("hydraulic_kW"),
        "Surface electrical (kW)": prod.get("surface_kW"),
        "Annual energy (MWh/yr)": prod.get("annual_MWh"),
    }
    st.markdown("**ESP — production** (hydraulic duty spec; not a selected pump model, no Hz)")
    show_df(pd.DataFrame(esp_tbl.items(), columns=["parameter", "value"]))
    inj_tbl = {
        "Injection well datum (m)": v3.get("inj_depth_m"),
        "Required injection BHP (bar)": v3.get("inj_bhp_bar"),
        "Total dynamic head (m)": inj.get("TDH_m"),
        "Pump needed": inj.get("pump_needed"),
        "Hydraulic power (kW)": inj.get("hydraulic_kW"),
        "Surface electrical (kW)": inj.get("surface_kW"),
        "Annual energy (MWh/yr)": inj.get("annual_MWh"),
    }
    st.markdown("**Injection pump (duty)**")
    show_df(pd.DataFrame(inj_tbl.items(), columns=["parameter", "value"]))
    st.markdown("**DHS circulator**")
    show_df(pd.DataFrame({"parameter": ["Circulator electrical (kW)"], "value": [eng.get("circ_kW")]}))
    st.markdown("**IPR / VLP ESP**")
    try:
        comm = build_commingled(RES_CFG)
        png, iprrows = AF.ipr_vlp(comm, v3, cfg, tmp, esp_pipe_id_m=getattr(ESPMOD,'PROD_PIPE_ID_M',0.1016))
        c1, c2 = st.columns([1.2, 1])
        c1.image(png, width=DISPLAY_W)
        c2.dataframe(_fmt_frame(pd.DataFrame(iprrows)), use_container_width=True, hide_index=True, height=420)
    except Exception as e:
        st.warning(f"IPR/VLP skipped: {e}")

# ================================== DHS ====================================
with tab_dhs:
    st.subheader("District-heating system — sized PHE, pipeline & injection")
    Rd = eng.get("_dhs", {}) or {}
    phe = Rd.get("PHE", {}) if isinstance(Rd.get("PHE"), dict) else {}
    area_req = phe.get("A_req_m2", eng.get("phe_area_m2"))
    area_tot = phe.get("A_total_m2", eng.get("phe_area_m2"))
    dhs_tbl = {
        "PHE type": phe.get("type"),
        "Plate class": phe.get("plate_class"),
        "Single-plate area (m²)": phe.get("A_plate_m2"),
        "Required PHE area, A_req (m²)": area_req,
        "Selected/installed PHE area, A_total (m²)": area_tot,
        "Overall U (W/m²K)": phe.get("U_Wm2K"),
        "LMTD (K)": phe.get("LMTD_K"),
        "Approach, hot end (K)": Rd.get("approach_hot"),
        "Approach, cold end (K)": Rd.get("approach_cold"),
        "Duty Q_HE (kW)": eng.get("Q_HE_kW"),
        "Brine in (°C)": Rd.get("T_brine_in", v3.get("wellhead_T_C")),
        "Brine out → injection (°C)": Rd.get("T_brine_out", eng.get("brine_out_C")),
        "DH return into PHE (°C)": Rd.get("T_HE_in"),
        "DH SUPPLY from PHE (°C)": Rd.get("T_HE_out"),
        "DH supply at city (°C)": Rd.get("T_DHS_in"),
        "DH return from city (°C)": Rd.get("T_DHS_out"),
        "Delivered heat (kW)": eng.get("delivered_kW"),
        "Pipeline heat loss (kW)": eng.get("pipe_loss_kW"),
        "DH carrier Ø, inner (mm)": cfg.dh_pipe_DN_m * 1000,
        "Injection line Ø, inner (mm)": cfg.brine_pipe_DN_m * 1000,
        "Circulator ΔP (bar)": (Rd.get("dp_pump_Pa", 0.0) or 0.0) / 1e5,
        "Circulator electrical (kW)": eng.get("circ_kW"),
    }
    show_df(pd.DataFrame([(k, v) for k, v in dhs_tbl.items() if v is not None],
            columns=["parameter", "value"]))

    # --- pipeline hydraulics at the current (peak) flow: velocity + pressure drop ---
    H = eng.get("hydraulics", {})
    if H:
        bl, dh = H.get("brine_line", {}), H.get("dh_line", {})
        hyd_rows = [
            {"line": "Brine reinjection", "DN now (mm)": bl.get("D_m", 0)*1000,
             "length (m)": bl.get("L_m"), "velocity (m/s)": bl.get("v_ms"),
             "Δp (bar)": bl.get("dP_bar"),
             "min DN for ≤2.5 m/s (mm)": bl.get("DN_min_m", 0)*1000},
            {"line": "DH carrier (supply+return)", "DN now (mm)": dh.get("D_m", 0)*1000,
             "length (m)": (dh.get("L_one_way_m") or 0)*2, "velocity (m/s)": dh.get("v_ms"),
             "Δp (bar)": dh.get("dP_pipe_bar"),
             "min DN for ≤2.5 m/s (mm)": dh.get("DN_min_m", 0)*1000},
        ]
        st.markdown("**Pipeline hydraulics (at peak/design flow)**")
        show_df(pd.DataFrame(hyd_rows))
        st.caption(f"DH circulator head (pipe supply+return + PHE) = "
                   f"{dh.get('dP_circuit_bar', 0):.2f} bar. Brine-line Δp is added at the injection "
                   f"end. Velocity ≤ ~2.5 m/s and DH gradient ≤ ~150 Pa/m are the comfort guidelines.")
        for w in H.get("warnings", []):
            st.warning("⚠ " + w)
        if not H.get("warnings"):
            st.success("Pipe velocities and DH pressure gradient are within comfortable design limits at this flow.")

    try:
        st.image(GHE.draw_system_scheme(eng["_dhs"], tmp), width=int(DISPLAY_W*1.6),
                 caption="DH system schematic — brine, PHE, pipeline & DH station (live results)")
    except Exception as e:
        st.warning(f"System schematic skipped: {e}")

# =========================== ECONOMY & FINANCE =============================
with tab_eco:
    st.subheader("Headline results")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Installed", f"{ener['installed_kWth']/1000:.2f} MWth")
    c2.metric("LCOH", f"€{cf['lcoh']:,.1f}/MWhth")
    c3.metric("Project NPV", f"€{cf['proj_npv']:,.0f}")
    irr = cf["proj_irr"]; c4.metric("Project IRR", "n/a" if irr != irr else f"{irr*100:.1f}%")
    c5, c6, c7, c8 = st.columns(4)
    c5.metric("Delivered heat (yr1)", f"{ener['delivered_MWh_y0']:,.0f} MWh")
    c6.metric("Discounted payback", "—" if cf["disc_payback"] is None else f"{cf['disc_payback']:.0f} yr")
    dscr = cf["dscr_min"]; c7.metric("Min DSCR", "∞" if dscr == float("inf") else f"{dscr:.2f}")
    c8.metric("Total CAPEX", f"€{capex_total:,.0f}")

    colA, colB = st.columns(2)
    with colA:
        st.markdown("**CAPEX breakdown (€)**")
        dfc = pd.DataFrame([(k, v) for k, v in clean(cap).items() if isinstance(v, (int, float))],
                           columns=["item", "EUR"])
        show_df(dfc)
        from matplotlib.ticker import FuncFormatter
        fig = matplotlib.figure.Figure(figsize=(8.5, 3.8), dpi=140); ax = fig.subplots()
        ax.barh(dfc["item"], dfc["EUR"], color="#1f4e79"); ax.invert_yaxis()
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, _: f"{x:,.0f}"))
        ax.set_xlabel("EUR"); ax.set_title("CAPEX breakdown"); fig.tight_layout()
        cp = os.path.join(tmp, "capex_bar.png"); fig.savefig(cp)
        st.image(cp, width=int(DISPLAY_W*1.4))
    with colB:
        st.markdown("**OPEX breakdown, yr1 (€/yr)**")
        dfo = pd.DataFrame([(k, v) for k, v in clean(op).items() if isinstance(v, (int, float))],
                           columns=["item", "EUR/yr"])
        show_df(dfo)

    st.markdown("**Energy profile & doublet decline**")
    yrs = np.arange(1, cfg.project_life_yr + 1)
    inst_MW = ener["installed_kWth"] / 1000.0 * ener["decline"]
    ep = pd.DataFrame({"year": yrs,
                       "Installed power (MW)": np.round(inst_MW, 3),
                       "Delivered heat (MWh)": np.round(ener["delivered_MWh"], 0),
                       "Decline factor (–)": np.round(ener["decline"], 3)})
    show_df(ep, height=240)

    st.markdown("**30-year cash flow**")
    if cf.get("rows"):
        show_df(pd.DataFrame(cf["rows"]), height=360)

    st.markdown("**Figures**")
    pay = next((f for f in figs if "payback" in f), None)
    if pay and os.path.exists(pay):
        st.image(pay, width=DISPLAY_W, caption="Cumulative discounted cash flow")
    try:
        tor = M.tornado(R, tmp)
        if os.path.exists(tor):
            st.image(tor, width=DISPLAY_W, caption="NPV sensitivity (tornado)")
    except Exception:
        pass

    try:
        st.download_button("⬇ Download results (Excel)", build_excel(R),
                           file_name="Ze1_results.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        st.caption(f"(Excel export unavailable: {e})")
