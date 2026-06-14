# NOTE: prices/OPEX/CO2 in this file are NOT used by the economic model.
# ECONOMICS HANDLED IN economy*.py (this script supplies physical sizing only).
"""
================================================================================
GEOTHERMAL DH — HEAT EXCHANGER & SURFACE EQUIPMENT (SENSITIVITY ANALYSIS)
================================================================================
This is the sensitivity-analysis variant of geothermal_HE.py.  It evaluates up
to three DH-network design scenarios in one run:

    Option A : existing 2G/3G DH network          (T_DHS_out typ. 60 °C)
    Option B : upgraded low-T DH network          (T_DHS_out typ. 40 °C)
               for new low-energy buildings with floor heating / low-T radiators
    Option C : HP-assisted cascaded DH network    (T_DHS_out typ. 15 °C)
               return from a third stage with consumer-side heat pumps;
               the HP itself is OUT OF SCOPE for this script — only the HE,
               pipeline and circulation pump are computed down to 15 °C.

Each option computes the same outputs as geothermal_HE.py:
    T_HE_out, T_DHS_in, T_HE_in, T_brine_out, m_water, Q_HE, pipe ΔP,
    PHE and S&T sizing, HE-type recommendation, pump electrical power,
    annual energy and OPEX, plus PHE and system schematics.

Plus the following sensitivity-specific features:

  1. AUTO-APPROACH SELECTION
     Hot- and cold-end approach temperatures are no longer fixed by the user.
     They are auto-selected from published industry bands based on T_brine_in
     and T_DHS_out:
        Hot end (T_brine_in − T_HE_out):
           T_brine_in ≥ 90 °C      → 5 K
           70 °C ≤ T_brine_in < 90 → 4 K
           T_brine_in < 70 °C      → 3 K
        Cold end (T_HE_in − T_brine_out):
           T_DHS_out ≥ 55 °C       → 5 K
           40 °C ≤ T_DHS_out < 55  → 4 K
           25 °C ≤ T_DHS_out < 40  → 3 K
           T_DHS_out < 25 °C       → 3 K + warning (manufacturer quote needed)
     Sources for these bands:  Alfa Laval (2024) Industrial Line GPHE clean
     water-to-water service; Kakaç et al. (2020) Heat Exchangers Tab. 7.4;
     Mertoglu (2001) Turkish geothermal DH experience (Kirsehir 54 °C case);
     Lund et al. (2014) 4GDH review.

  2. MANUAL OVERRIDE
     Each option's approach can be manually overridden via
     approach_hot_X / approach_cold_X.  Set to None → auto-select.
     Set to a number → use that number; auto-default is still reported for
     comparison; warning fires if outside 3-8 K published range.

  3. AUTO-FALLBACK
     If the auto-selected design is physically infeasible (LMTD < 0.5 K,
     PHE A > 2000 m², plate count > 2000, U < 400 W/(m²·K), F-correction
     fails, etc.), the script loosens the cold approach by 1 K and retries,
     up to a 6 K ceiling.  All retries are logged in console + Excel.
     The fallback NEVER fires when approaches are manually overridden.

  4. SANITY WARNINGS
     Borderline designs (area > 800 m², plates > 800, U < 800, NTU > 8,
     ΔP > 60 kPa, etc.) print explicit warnings but the design is still
     reported.  This lets the paper show "we tried, here are the caveats".

  5. PER-OPTION TOGGLE
     Set T_DHS_out_X = None (or 0) to skip Option X entirely.  Useful when
     a given site does not have low-T or HP-assisted future stages.

  6. PARAMETRIC SENSITIVITY SWEEP
     A continuous sweep of T_DHS_out from 15 °C to 65 °C in 5 °C steps
     produces a 2×2 figure showing Q_HE, m_water, A_PHE and OPEX as
     functions of T_DHS_out, with the chosen A/B/C operating points marked.

OUTPUTS:
    Console : full per-option output + final comparison table
    Excel   : one sheet per enabled option (Option A / B / C)
              + Sensitivity summary sheet
    PNGs    : PHE_scheme_A.png, system_scheme_A.png, ... (one pair per option)
              sensitivity_plot.png

================================================================================
References (Harvard) — additions vs. geothermal_HE.py:
  Lund, H. et al. (2014) 'Fourth-Generation District Heating (4GDH)',
      Energy 68, 1-11.  doi:10.1016/j.energy.2014.02.089
  Mertoglu, O. (2001) 'Geothermal District Heating Experience in Turkey',
      Proc. International Geothermal Conference, Reykjavik 2001.
  Persson, U. & Werner, S. (2011) 'Heat distribution and the future
      competitiveness of district heating', Applied Energy 88(3), 568-576.
================================================================================
"""

import math
import os

# ==============================================================================
# PART 1 — INPUT PARAMETERS  (edit only this block)
# ==============================================================================

# ---- 1.1 Brine side (COMMON to all options) -------------------------------
T_brine_in   = 94.0    # brine wellhead T = HE primary inlet [°C]
Q_brine      = 35.0    # brine volumetric flow at wellhead [L/s]
                       # (treated as ≈ kg/s; salinity disregarded per user choice)

# ---- 1.2 Scenario block ---------------------------------------------------
# Each option has three inputs:
#   T_DHS_out_X     — DH-return T entering HE secondary [°C]
#                     Set to None (or 0) to SKIP this option entirely.
#   approach_hot_X  — hot-end approach (T_brine_in − T_HE_out) [K]
#                     None → auto-select from T_brine_in band (see Part 8)
#                     <number> → manually use this value (warning if outside 3-8 K)
#   approach_cold_X — cold-end approach (T_HE_in − T_brine_out) [K]
#                     None → auto-select from T_DHS_out band
#                     <number> → manually use this value
# ---------------------------------------------------------------------------

# Option A — existing 2G/3G DH network (high-T radiators, conventional return)
T_DHS_out_A      = 60.0       # set to None to skip Option A
approach_hot_A   = None       # None = auto
approach_cold_A  = None       # None = auto
PHE_plate_class_A = "auto"    # "auto" = pick smallest fitting class;
                              # or "M10" / "M15" / "T20" / "MX25" / "M30"

# Option B — upgraded low-T DH for new low-energy buildings (floor heat etc.)
T_DHS_out_B      = 40.0
approach_hot_B   = None
approach_cold_B  = None
PHE_plate_class_B = "auto"

# Option C — HP-assisted cascaded DH (HP itself out of scope for now)
T_DHS_out_C      = 15.0
approach_hot_C   = None
approach_cold_C  = None
PHE_plate_class_C = "auto"

# ---- 1.3 Ground / ambient (COMMON) ----------------------------------------
T_soil_january = 6.2   # January ground temperature at 1 m [°C]
                       # measured value at the design site (use local
                       # meteorological-service ground-temperature data).

# ---- 1.4 Pipeline geometry (COMMON) ---------------------------------------
L_pipe       = 2345.0  # one-way pipe length, HE → DH station [m]
burial_depth = 1.0     # depth from ground surface to TOP of casing [m]

# Logstor / Kingspan EN 253 single pipe, INSULATION SERIES 2.
D_inner       = 0.2101    # DN 200, carrier inner diameter [m]
D_outer_steel = 0.2191    # carrier outer diameter [m]  (P235GH per EN 10217)
t_insulation  = 0.0639    # PUR foam thickness [m]   ((0.355 − 0.219 − 0.009)/2)
t_casing      = 0.0045    # HDPE jacket thickness [m]
roughness     = 0.045e-3  # absolute roughness, new commercial steel [m]
k_steel = 50.0; k_PUR = 0.027; k_HDPE = 0.43      # [W/(m·K)]
k_soil  = 1.60                                     # moist clay/loam (Pannonian)

# Fittings (per direction, scaled for ≈ 2.3 km buried trunk):
N_elbow_90    = 78
N_elbow_45    = 31
N_gate_valve  = 5
N_tee         = 3
N_reducer     = 2
N_check_valve = 1
N_strainer    = 1
N_flow_meter  = 1

# ---- 1.5 PHE — gasketed plate-and-frame, AISI 316L SS ---------------------
# Plate dimensions and per-frame limits are taken from the Alfa Laval
# Industrial Line (M10, M15, T20, MX25, M30 — published technical datasheets,
# directindustry.com 2024).  "max_plates" is the conservative typical
# single-frame limit (the technical maximum can be 900-1000 in some larger
# frames, but 700 is the safe design target).  "max_area_single_frame" is
# computed as max_plates × A_plate (the area covered before splitting into
# multiple parallel frames is needed).
PHE_CLASSES = {
    "M10":  dict(W=0.470, H=1.084, A_plate=0.510, max_plates=700, max_flow_kg_s=50,  max_area_single_frame=357),
    "M15":  dict(W=0.610, H=1.885, A_plate=1.150, max_plates=700, max_flow_kg_s=80,  max_area_single_frame=805),
    "T20":  dict(W=0.780, H=2.100, A_plate=1.638, max_plates=700, max_flow_kg_s=180, max_area_single_frame=1147),
    "MX25": dict(W=0.920, H=2.895, A_plate=2.663, max_plates=700, max_flow_kg_s=250, max_area_single_frame=1864),
    "M30":  dict(W=1.150, H=2.882, A_plate=3.314, max_plates=700, max_flow_kg_s=650, max_area_single_frame=2320),
}
# Order from smallest to largest (used by auto-pick logic)
PHE_CLASS_ORDER = ["M10", "M15", "T20", "MX25", "M30"]
# Fraction of max_area_single_frame above which auto-pick steps up to next
# class — provides headroom for fouling allowance and future capacity growth.
PHE_AUTO_HEADROOM = 0.85

PHE_chevron_deg          = 60.0
PHE_plate_thickness_m    = 0.0005
PHE_plate_gap_m          = 0.003
PHE_fouling_factor_m2KW  = 8.6e-5
k_plate_WmK              = 16.0

# ---- 1.6 Shell-and-tube — TEMA E, 1 shell + 2 tube passes -----------------
ST_tube_OD               = 0.01905
ST_tube_wall             = 0.00165
ST_tube_length           = 4.88
ST_pitch_ratio           = 1.25
ST_layout                = "triangular"
ST_n_tube_passes         = 2
ST_baffle_cut            = 0.25
ST_baffle_pitch_factor   = 0.40
ST_fouling_factor_m2KW   = 1.76e-4
k_tube_WmK               = 16.0

# ---- 1.7 Pump and drive (COMMON) ------------------------------------------
eta_pump  = 0.75
eta_motor = 0.95
eta_VFD   = 0.97

# ---- 1.8 Operating hours (COMMON) -----------------------------------------
FLH = 1800.0          # full-load hours per year

# ---- 1.9 Electricity tariff (COMMON) --------------------------------------
ELEC_TARIFF_EUR_kWh_excl_VAT = 0.1763   # Eurostat nrg_pc_205 Croatia H1 2025
VAT_RATE                     = 0.25     # Croatian standard VAT 25 %

# ---- 1.10 Numerics & output -----------------------------------------------
ITER_TOL_KGS = 1e-3
MAX_ITER     = 40
VERBOSE      = False
SAVE_EXCEL          = False
SAVE_PHE_SCHEMES    = False
SAVE_SYSTEM_SCHEMES = False
SAVE_SENSITIVITY_PLOT = False

# Sensitivity-sweep range (for the 2×2 plot)
SWEEP_T_DHS_OUT_MIN = 15.0
SWEEP_T_DHS_OUT_MAX = 65.0
SWEEP_T_DHS_OUT_STEP = 5.0


# ==============================================================================
# PART 2 — WATER PROPERTIES (pure water; salinity disregarded)
# ==============================================================================
GRAV = 9.80665

def cp_water(T_C, p_bar=4.0):
    return 4217.7 - 3.166*T_C + 0.05010*T_C**2 - 3.254e-4*T_C**3 + 1.014e-6*T_C**4

def rho_water(T_C, p_bar=4.0):
    a = (999.83952 + 16.945176*T_C - 7.9870401e-3*T_C**2
         - 46.170461e-6*T_C**3 + 105.56302e-9*T_C**4 - 280.54253e-12*T_C**5)
    b = 1.0 + 16.879850e-3*T_C
    return a/b

def mu_water(T_C, p_bar=4.0):
    T_K = T_C + 273.15
    return 1.0e-3 * math.exp(-3.7188 + 578.919/(T_K - 137.546))

def k_water(T_C, p_bar=4.0):
    return 0.5694 + 1.494e-3*T_C - 6.71e-6*T_C**2

def Pr_water(T_C, p_bar=4.0):
    return cp_water(T_C, p_bar) * mu_water(T_C, p_bar) / k_water(T_C, p_bar)

def water_props(T_C, p_bar=4.0):
    return dict(rho=rho_water(T_C,p_bar), mu=mu_water(T_C,p_bar),
                cp=cp_water(T_C,p_bar), k=k_water(T_C,p_bar), Pr=Pr_water(T_C,p_bar))


# ==============================================================================
# PART 3 — HYDRAULICS
# ==============================================================================
def colebrook(Re, eps_D, tol=1e-9, itmax=60):
    if Re < 2300.0:
        return 64.0/Re
    f = 0.25 / (math.log10(eps_D/3.7 + 5.74/Re**0.9))**2
    for _ in range(itmax):
        f_new = 1.0 / (-2.0*math.log10(eps_D/3.7 + 2.51/(Re*math.sqrt(f))))**2
        if abs(f_new - f) < tol: return f_new
        f = f_new
    return f

def K_fittings():
    K = {"elbow_90":0.30, "elbow_45":0.20, "gate_valve":0.15, "tee":0.90,
         "reducer":0.20, "check_valve":2.00, "strainer":2.50, "flow_meter":0.50}
    counts = {"elbow_90":N_elbow_90, "elbow_45":N_elbow_45,
              "gate_valve":N_gate_valve, "tee":N_tee, "reducer":N_reducer,
              "check_valve":N_check_valve, "strainer":N_strainer,
              "flow_meter":N_flow_meter}
    return sum(counts[n]*K[n] for n in K), counts, K

def pipe_hydraulics(m_dot, props, D, L, eps, sumK):
    A = math.pi*D**2/4.0
    v  = m_dot/(props["rho"]*A)
    Re = props["rho"]*v*D/props["mu"]
    f  = colebrook(Re, eps/D)
    dpf = f*(L/D)*0.5*props["rho"]*v**2
    dpm = sumK*0.5*props["rho"]*v**2
    return dict(v=v, Re=Re, f=f, dp_friction=dpf, dp_minor=dpm, dp_total=dpf+dpm)


# ==============================================================================
# PART 4 — INTERNAL HTC & PIPELINE UA
# ==============================================================================
def h_internal_Gnielinski(Re, Pr, k, D, f):
    if Re < 2300:   Nu = 3.66
    elif Re < 3000:
        Nu_t = (f/8.0)*(Re-1000.0)*Pr / (1.0 + 12.7*math.sqrt(f/8.0)*(Pr**(2/3)-1.0))
        x = (Re-2300.0)/700.0
        Nu = 3.66*(1-x) + Nu_t*x
    else:
        Nu = (f/8.0)*(Re-1000.0)*Pr / (1.0 + 12.7*math.sqrt(f/8.0)*(Pr**(2/3)-1.0))
    return Nu*k/D

def UA_buried_pipe(L, D_i, D_st, t_ins, t_cas, k_st, k_in, k_cas, k_so,
                   H_top, h_i):
    D_steel = D_st
    D_ins   = D_st + 2*t_ins
    D_cas   = D_ins + 2*t_cas
    H_c     = H_top + D_cas/2.0
    R_conv = 1.0/(h_i*math.pi*D_i*L)
    R_st   = math.log(D_steel/D_i)/(2*math.pi*k_st*L)
    R_in   = math.log(D_ins/D_steel)/(2*math.pi*k_in*L)
    R_cas  = math.log(D_cas/D_ins)/(2*math.pi*k_cas*L)
    R_so   = math.acosh(2*H_c/D_cas)/(2*math.pi*k_so*L)
    R_tot  = R_conv + R_st + R_in + R_cas + R_so
    return 1.0/R_tot, dict(R_conv=R_conv, R_steel=R_st, R_ins=R_in,
                           R_cas=R_cas, R_soil=R_so, R_total=R_tot)

def T_out_pipe(T_in, T_amb, UA, m, cp):
    return T_amb + (T_in - T_amb)*math.exp(-UA/(m*cp))


# ==============================================================================
# PART 5 — PLATE HEAT EXCHANGER (Martin 1996)
# ==============================================================================
def LMTD_counter(T_hi, T_ho, T_ci, T_co):
    d1, d2 = T_hi-T_co, T_ho-T_ci
    if abs(d1-d2) < 1e-6: return 0.5*(d1+d2)
    if d1*d2 <= 0: return float("nan")
    return (d1-d2)/math.log(d1/d2)

def PHE_friction_Martin(Re, beta_deg):
    beta = math.radians(beta_deg)
    if Re < 2000:
        f0, f1 = 64.0/Re, 597.0/Re + 3.85
    else:
        f0 = (1.8*math.log10(Re)-1.5)**(-2)
        f1 = 39.0/Re**0.289
    inv = (math.cos(beta)
           /math.sqrt(0.18*math.tan(beta)+0.36*math.sin(beta)+f0/math.cos(beta))
           + (1.0-math.cos(beta))/math.sqrt(3.8*f1))
    return 1.0/inv**2

def PHE_Nu_Martin(Re, Pr, beta_deg, f):
    beta = math.radians(beta_deg)
    return 0.205*Pr**(1.0/3.0)*(f*Re**2*math.sin(2*beta))**0.374

def size_PHE(m_h, props_h, T_hi, T_ho, m_c, props_c, T_ci, T_co, plate_class="M10"):
    """Size a plate-and-frame HE for a specific Alfa Laval class.
    The Martin (1996) correlations depend on plate gap and chevron angle
    (constant across the Industrial Line) and on hydraulic Re, which is
    affected by plate width via the channel cross-section.  Changing class
    therefore changes U mildly (typically ±10-20 %) and N_plates strongly."""
    if plate_class not in PHE_CLASSES:
        raise ValueError(f"Unknown PHE plate class: {plate_class!r}. "
                         f"Valid: {list(PHE_CLASSES)}")
    pc = PHE_CLASSES[plate_class]
    plate_W = pc["W"]; plate_H = pc["H"]; A_plate = pc["A_plate"]
    if (T_hi-T_co) <= 0 or (T_ho-T_ci) <= 0:
        raise ValueError(f"PHE crossover: counter-current infeasible "
                         f"(T_hi={T_hi:.2f}, T_co={T_co:.2f}, "
                         f"T_ho={T_ho:.2f}, T_ci={T_ci:.2f}).")
    Q  = m_h*props_h["cp"]*(T_hi-T_ho)
    Qc = m_c*props_c["cp"]*(T_co-T_ci)
    lmtd = LMTD_counter(T_hi, T_ho, T_ci, T_co)
    Dh = 2.0*PHE_plate_gap_m
    A_ch = plate_W*PHE_plate_gap_m
    foul = PHE_fouling_factor_m2KW

    def one_pass(N):
        G_h = (m_h/N)/A_ch; G_c = (m_c/N)/A_ch
        Re_h, Re_c = G_h*Dh/props_h["mu"], G_c*Dh/props_c["mu"]
        f_h = PHE_friction_Martin(Re_h, PHE_chevron_deg)
        f_c = PHE_friction_Martin(Re_c, PHE_chevron_deg)
        Nu_h = PHE_Nu_Martin(Re_h, props_h["Pr"], PHE_chevron_deg, f_h)
        Nu_c = PHE_Nu_Martin(Re_c, props_c["Pr"], PHE_chevron_deg, f_c)
        h_h, h_c = Nu_h*props_h["k"]/Dh, Nu_c*props_c["k"]/Dh
        R_plate = PHE_plate_thickness_m/k_plate_WmK
        U = 1.0/(1.0/h_h + foul + R_plate + foul + 1.0/h_c)
        dp_h = f_h*(plate_H/Dh)*G_h**2/(2.0*props_h["rho"])
        dp_c = f_c*(plate_H/Dh)*G_c**2/(2.0*props_c["rho"])
        return Re_h, Re_c, h_h, h_c, U, dp_h, dp_c

    N = 50
    for _ in range(50):
        Re_h, Re_c, h_h, h_c, U, dp_h, dp_c = one_pass(N)
        A_req = Q/(U*lmtd)
        N_new = max(2, math.ceil(A_req/(2.0*A_plate)))
        if N_new == N: break
        N = N_new
    A_total = N*2*A_plate
    C_h, C_c = m_h*props_h["cp"], m_c*props_c["cp"]
    C_min, C_max = min(C_h,C_c), max(C_h,C_c)
    Cr = C_min/C_max
    NTU = U*A_total/C_min
    eps = (NTU/(1.0+NTU) if abs(Cr-1.0)<1e-6
           else (1.0-math.exp(-NTU*(1.0-Cr)))/(1.0-Cr*math.exp(-NTU*(1.0-Cr))))

    foot_W = plate_W + 0.30
    foot_H = plate_H + 0.45
    foot_L = (2*N+1)*0.004 + 0.60
    weight_kg = (2*N+1)*A_plate*7900*PHE_plate_thickness_m + 250.0

    # Single-frame fit assessment
    N_plates = 2*N+1
    fits_single = (N_plates <= pc["max_plates"]) and (A_total <= pc["max_area_single_frame"])
    N_frames_needed = max(1, math.ceil(A_total / pc["max_area_single_frame"]))

    return dict(type="PHE (gasketed)", plate_class=plate_class,
                plate_W_m=plate_W, plate_H_m=plate_H, A_plate_m2=A_plate,
                max_plates_per_frame=pc["max_plates"],
                max_area_single_frame_m2=pc["max_area_single_frame"],
                Q_W=Q, Q_check_W=Qc, LMTD_K=lmtd, U_Wm2K=U,
                A_req_m2=A_req, A_total_m2=A_total, N_pairs=N, N_plates=N_plates,
                fits_single_frame=fits_single, N_frames_needed=N_frames_needed,
                NTU=NTU, effectiveness=eps,
                Re_h=Re_h, Re_c=Re_c, h_h=h_h, h_c=h_c,
                dp_brine_Pa=dp_h, dp_water_Pa=dp_c,
                approach_hot=T_hi-T_co, approach_cold=T_ho-T_ci,
                footprint_LWH_m=(foot_L,foot_W,foot_H),
                weight_kg=weight_kg)


def auto_pick_PHE_class(m_h, props_h, T_hi, T_ho, m_c, props_c, T_ci, T_co):
    """Iterate from M10 upward; return the smallest class that fits in a
    single frame at PHE_AUTO_HEADROOM (default 85 %) of capacity.  If even
    M30 cannot fit at 85 %, return M30 with N_frames_needed > 1."""
    for cls in PHE_CLASS_ORDER:
        r = size_PHE(m_h, props_h, T_hi, T_ho, m_c, props_c, T_ci, T_co, cls)
        max_pl  = PHE_CLASSES[cls]["max_plates"]
        max_A   = PHE_CLASSES[cls]["max_area_single_frame"]
        if (r["N_plates"] <= PHE_AUTO_HEADROOM*max_pl
                and r["A_total_m2"] <= PHE_AUTO_HEADROOM*max_A):
            return r
    # Even M30 doesn't fit at 85 % — return M30 with multi-frame flag set
    return size_PHE(m_h, props_h, T_hi, T_ho, m_c, props_c, T_ci, T_co, "M30")


def resolve_plate_class(plate_class_input, m_h, props_h, T_hi, T_ho,
                        m_c, props_c, T_ci, T_co):
    """Resolve PHE_plate_class input (string).
       'auto' / None  → call auto-pick.
       'M10' .. 'M30' → use that class directly.
       Returns (PHE_result_dict, notes_list)."""
    notes = []
    if plate_class_input is None or str(plate_class_input).lower() == "auto":
        r = auto_pick_PHE_class(m_h, props_h, T_hi, T_ho, m_c, props_c, T_ci, T_co)
        notes.append(f"PHE PLATE CLASS: {r['plate_class']} (auto-picked: smallest "
                     f"Alfa Laval Industrial Line class whose single-frame "
                     f"capacity at {PHE_AUTO_HEADROOM*100:.0f} % headroom covers "
                     f"A_required = {r['A_req_m2']:.0f} m²)")
    else:
        cls = str(plate_class_input).upper()
        if cls not in PHE_CLASSES:
            raise ValueError(f"PHE_plate_class = {plate_class_input!r} is not "
                             f"valid. Valid values: 'auto', "
                             f"{', '.join(repr(c) for c in PHE_CLASS_ORDER)}")
        r = size_PHE(m_h, props_h, T_hi, T_ho, m_c, props_c, T_ci, T_co, cls)
        # Compare against what auto would have picked
        r_auto = auto_pick_PHE_class(m_h, props_h, T_hi, T_ho, m_c, props_c, T_ci, T_co)
        notes.append(f"PHE PLATE CLASS: {cls} (MANUALLY SET; auto-pick would "
                     f"have chosen {r_auto['plate_class']})")
        if not r["fits_single_frame"]:
            notes.append(f"  Manual choice {cls} does not fit single frame: "
                         f"requires {r['N_frames_needed']} parallel frames "
                         f"of {PHE_CLASSES[cls]['max_plates']} plates each.")
    return r, notes


# ==============================================================================
# PART 6 — SHELL-AND-TUBE (Kern + fixed v_t design constraint)
# ==============================================================================
def F_LMTD_1shell_2tube(P, R):
    if abs(R-1.0) < 1e-6:
        num = P*math.sqrt(2.0)/(1.0-P)
        den = math.log((2.0-P*(2.0-math.sqrt(2.0)))/(2.0-P*(2.0+math.sqrt(2.0))))
        return num/den
    s = math.sqrt(R**2+1.0)
    num = s*math.log((1.0-P)/(1.0-P*R))
    den = (R-1.0)*math.log((2.0-P*(R+1.0-s))/(2.0-P*(R+1.0+s)))
    return num/den

def size_ST(m_h, props_h, T_hi, T_ho, m_c, props_c, T_ci, T_co):
    if (T_hi-T_co) <= 0 or (T_ho-T_ci) <= 0:
        raise ValueError("S&T crossover: counter-current infeasible.")
    Q  = m_h*props_h["cp"]*(T_hi-T_ho)
    Qc = m_c*props_c["cp"]*(T_co-T_ci)
    lmtd = LMTD_counter(T_hi, T_ho, T_ci, T_co)

    P = (T_co-T_ci)/(T_hi-T_ci) if (T_hi-T_ci)!=0 else 0.0
    R = (T_hi-T_ho)/(T_co-T_ci) if (T_co-T_ci)!=0 else 1.0
    feasible_12 = True; F_note = ""
    try:
        F_calc = F_LMTD_1shell_2tube(P, R)
        if not math.isfinite(F_calc) or F_calc <= 0:
            raise ValueError("F undefined")
        if F_calc < 0.75:
            feasible_12 = False
            F_note = (f"F = {F_calc:.3f} < 0.75 (TEMA limit) — 1-2 S&T uneconomic "
                      f"at this close approach. P={P:.3f}, R={R:.3f}.")
            F = F_calc
        else:
            F = F_calc
    except (ValueError, ZeroDivisionError):
        feasible_12 = False
        F_note = (f"1-2 TEMA E shell mathematically infeasible at close approach "
                  f"(P={P:.3f}, R={R:.3f}). Pure counter-current 1-1 needed.")
        F = 1.0

    D_o = ST_tube_OD; D_i = ST_tube_OD - 2*ST_tube_wall; P_t = ST_pitch_ratio*D_o
    v_t_design = 1.5
    A_tube_in = math.pi*D_i**2/4.0
    N_tubes = max(20, math.ceil(m_h*ST_n_tube_passes
                                / (props_h["rho"]*A_tube_in*v_t_design)))
    m_per_tube = m_h*ST_n_tube_passes/N_tubes
    v_t = m_per_tube/(props_h["rho"]*A_tube_in)
    Re_t = props_h["rho"]*v_t*D_i/props_h["mu"]
    f_t = colebrook(Re_t, 0.000046/D_i)
    h_t = h_internal_Gnielinski(Re_t, props_h["Pr"], props_h["k"], D_i, f_t)

    K1, n1 = (0.319, 2.142) if ST_layout == "triangular" else (0.215, 2.207)
    D_b = D_o*(N_tubes/K1)**(1.0/n1)
    D_s = D_b + 0.050
    B = ST_baffle_pitch_factor*D_s
    if ST_layout == "triangular":
        D_e = (1.10/D_o)*(P_t**2 - 0.917*D_o**2)
    else:
        D_e = (1.27/D_o)*(P_t**2 - 0.785*D_o**2)
    A_s = D_s*(P_t-D_o)*B/P_t
    G_s = m_c/A_s
    Re_s = G_s*D_e/props_c["mu"]
    h_s = 0.36*(props_c["k"]/D_e)*(Re_s**0.55)*(props_c["Pr"]**(1.0/3.0))

    R_wall = D_o*math.log(D_o/D_i)/(2*k_tube_WmK)
    U = 1.0/(D_o/(D_i*h_t) + ST_fouling_factor_m2KW*D_o/D_i
             + R_wall + ST_fouling_factor_m2KW + 1.0/h_s)

    A_req = Q/(U*lmtd*F)
    L_t_required = A_req / (N_tubes*math.pi*D_o)
    L_t = max(ST_tube_length, L_t_required)
    A_inst = N_tubes*math.pi*D_o*L_t

    dp_t = (4*f_t*L_t/D_i + 4*ST_n_tube_passes)*0.5*props_h["rho"]*v_t**2
    f_s = math.exp(0.576 - 0.19*math.log(max(Re_s,1.0)))
    N_baffles = max(1, math.ceil(L_t/B) - 1)
    dp_s = f_s*G_s**2*(N_baffles+1)*D_s/(2.0*props_c["rho"]*D_e)

    C_h, C_c = m_h*props_h["cp"], m_c*props_c["cp"]
    C_min, C_max = min(C_h,C_c), max(C_h,C_c)
    Cr = C_min/C_max
    NTU = U*A_inst/C_min
    eps = (NTU/(1.0+NTU) if abs(Cr-1.0)<1e-6
           else (1.0-math.exp(-NTU*(1.0-Cr)))/(1.0-Cr*math.exp(-NTU*(1.0-Cr))))

    foot_L = L_t + 0.80; foot_W = D_s + 0.20; foot_H = D_s + 0.20
    tube_mass = N_tubes*math.pi*D_o*L_t*ST_tube_wall*7850
    shell_mass = math.pi*D_s*L_t*0.012*7850
    weight_kg = tube_mass + shell_mass + 500

    return dict(type="Shell-and-tube (TEMA E, 1-2)", Q_W=Q, Q_check_W=Qc,
                LMTD_K=lmtd, F_correction=F, U_Wm2K=U,
                A_req_m2=A_req, A_total_m2=A_inst, N_tubes=N_tubes,
                D_shell_m=D_s, baffle_spacing_m=B, N_baffles=N_baffles,
                L_tube_m=L_t,
                NTU=NTU, effectiveness=eps,
                Re_tube=Re_t, Re_shell=Re_s, h_tube=h_t, h_shell=h_s,
                dp_brine_Pa=dp_t, dp_water_Pa=dp_s,
                approach_hot=T_hi-T_co, approach_cold=T_ho-T_ci,
                footprint_LWH_m=(foot_L,foot_W,foot_H), weight_kg=weight_kg,
                feasible_12=feasible_12, F_note=F_note,
                P_factor=P, R_factor=R, v_tube=v_t)


# ==============================================================================
# PART 7 — HE TYPE RECOMMENDATION
# ==============================================================================
def recommend_HE_type(PHE_res, ST_res, T_max_C, p_max_bar, fluid_type, duty_MW):
    reasons = []
    if T_max_C > 180:
        reasons.append("T_max > 180 °C → gasketed PHE unsuitable. S&T recommended.")
        return "S&T", reasons
    if p_max_bar > 25:
        reasons.append("p_max > 25 bar → gasketed PHE unsuitable. S&T recommended.")
        return "S&T", reasons
    if fluid_type in ("two-phase","viscous","fouling"):
        reasons.append(f"Fluid type '{fluid_type}' → S&T preferred "
                       "(Sinnott & Towler 2020 Ch. 12).")
        return "S&T", reasons
    if not ST_res.get("feasible_12", True):
        reasons.append(ST_res.get("F_note","S&T 1-2 infeasible at this approach"))
    if ST_res.get("feasible_12", True):
        PHE_vol = (PHE_res["footprint_LWH_m"][0]*PHE_res["footprint_LWH_m"][1]
                   *PHE_res["footprint_LWH_m"][2])
        ST_vol  = (ST_res["footprint_LWH_m"][0]*ST_res["footprint_LWH_m"][1]
                   *ST_res["footprint_LWH_m"][2])
        ratio = ST_vol/PHE_vol
        reasons.append(f"Footprint ratio (S&T / PHE) ≈ {ratio:.1f}× — PHE typically "
                       "3-5× more compact (Kakaç 2020; Shah & Sekulic 2003).")
        reasons.append(f"U_PHE = {PHE_res['U_Wm2K']:.0f} vs U_S&T = "
                       f"{ST_res['U_Wm2K']:.0f} W/(m²·K) → PHE gives "
                       f"{PHE_res['U_Wm2K']/ST_res['U_Wm2K']:.1f}× higher U.")
    reasons.append("PHE plates can be added/removed for capacity change; S&T fixed "
                   "at manufacture (Sinnott & Towler 2020 §12.7).")
    return "PHE", reasons


# ==============================================================================
# PART 8 — AUTO-APPROACH SELECTION & MANUAL OVERRIDE
# ==============================================================================
def auto_approaches(T_brine_in_C, T_DHS_out_C):
    """Return (hot_K, cold_K, hot_note, cold_note) based on T bands.

    Hot end (T_brine_in − T_HE_out):
      ≥ 90 °C        → 5 K   Alfa Laval (2024) GPHE clean water service
      70 °C – 90 °C  → 4 K   Kakaç et al. (2020) Tab. 7.4
      < 70 °C        → 3 K   Mertoglu (2001) Kirsehir close-approach case

    Cold end (T_HE_in − T_brine_out):
      ≥ 55 °C        → 5 K   conventional 2G/3G, Mertoglu (2001)
      40 °C – 55 °C  → 4 K   low-T DH, Alfa Laval (2024)
      25 °C – 40 °C  → 3 K   4GDH cascade, Lund et al. (2014)
      < 25 °C        → 3 K + WARNING — at the manufacturer economic floor;
                              direct quote recommended.
    """
    # Hot end
    if T_brine_in_C >= 90:
        hot = 5.0
        hot_note = (f"hot = {hot:.1f} K (auto: T_brine_in = {T_brine_in_C:.1f} °C "
                    f"≥ 90 °C, Alfa Laval 2024 GPHE clean water service)")
    elif T_brine_in_C >= 70:
        hot = 4.0
        hot_note = (f"hot = {hot:.1f} K (auto: 70 ≤ T_brine_in = "
                    f"{T_brine_in_C:.1f} °C < 90, Kakaç et al. 2020 Tab. 7.4)")
    else:
        hot = 3.0
        hot_note = (f"hot = {hot:.1f} K (auto: T_brine_in = {T_brine_in_C:.1f} °C "
                    f"< 70 °C, Mertoglu 2001 Kirsehir close-approach case)")

    # Cold end
    if T_DHS_out_C >= 55:
        cold = 5.0
        cold_note = (f"cold = {cold:.1f} K (auto: T_DHS_out = {T_DHS_out_C:.1f} °C "
                     f"≥ 55 °C, conventional 2G/3G practice, Mertoglu 2001)")
    elif T_DHS_out_C >= 40:
        cold = 4.0
        cold_note = (f"cold = {cold:.1f} K (auto: 40 ≤ T_DHS_out = "
                     f"{T_DHS_out_C:.1f} °C < 55, low-T DH, Alfa Laval 2024)")
    elif T_DHS_out_C >= 25:
        cold = 3.0
        cold_note = (f"cold = {cold:.1f} K (auto: 25 ≤ T_DHS_out = "
                     f"{T_DHS_out_C:.1f} °C < 40, 4GDH cascade, Lund et al. 2014)")
    else:
        cold = 3.0
        cold_note = (f"cold = {cold:.1f} K (auto: T_DHS_out = {T_DHS_out_C:.1f} °C "
                     f"< 25 °C, AT MANUFACTURER FLOOR — direct quotation "
                     f"recommended; Alfa Laval Industrial Line 2024)")
    return hot, cold, hot_note, cold_note


def resolve_approaches(T_brine_in_C, T_DHS_out_C, hot_override, cold_override):
    """Combine auto-selection with optional manual override.
    Returns (hot_K, cold_K, notes_list).  notes_list contains human-readable
    strings describing what was chosen and why (used in console + Excel)."""
    auto_hot, auto_cold, auto_hot_note, auto_cold_note = \
        auto_approaches(T_brine_in_C, T_DHS_out_C)
    notes = []
    # Hot end
    if hot_override is None:
        hot = auto_hot
        notes.append("HOT APPROACH: " + auto_hot_note)
    else:
        hot = float(hot_override)
        notes.append(f"HOT APPROACH: {hot:.1f} K (MANUALLY SET; auto-default "
                     f"would be {auto_hot:.1f} K)")
        if hot < 3.0:
            notes.append(f"  WARNING: {hot:.1f} K is below 3 K economic floor "
                         f"(Alfa Laval 2024); manufacturer quote recommended.")
        elif hot > 8.0:
            notes.append(f"  NOTE: {hot:.1f} K is loose; design will be conservative.")
    # Cold end
    if cold_override is None:
        cold = auto_cold
        notes.append("COLD APPROACH: " + auto_cold_note)
    else:
        cold = float(cold_override)
        notes.append(f"COLD APPROACH: {cold:.1f} K (MANUALLY SET; auto-default "
                     f"would be {auto_cold:.1f} K)")
        if cold < 3.0:
            notes.append(f"  WARNING: {cold:.1f} K is below 3 K economic floor "
                         f"(Alfa Laval 2024); manufacturer quote recommended.")
        elif cold > 8.0:
            notes.append(f"  NOTE: {cold:.1f} K is loose; design will be conservative.")
    return hot, cold, notes


# ==============================================================================
# PART 9 — SANITY WARNINGS & INFEASIBILITY DETECTION
# ==============================================================================
def design_sanity_warnings(R):
    """Inspect a converged R dict and return a list of human-readable warning
    strings.  Always called.  Does NOT trigger auto-fallback by itself."""
    warns = []
    if R is None: return warns
    P = R.get("PHE")
    if P is None: return warns
    A = P["A_total_m2"]; Np = P["N_plates"]; U = P["U_Wm2K"]
    L = P["LMTD_K"];     NTU = P["NTU"];    dpw = P["dp_water_Pa"]
    eps = P["effectiveness"]
    cls = P["plate_class"]
    max_pl = P["max_plates_per_frame"]
    max_A  = P["max_area_single_frame_m2"]

    # Multi-frame check
    if not P["fits_single_frame"]:
        N_fr = P["N_frames_needed"]
        # Find next larger class for the alternative recommendation
        idx = PHE_CLASS_ORDER.index(cls) if cls in PHE_CLASS_ORDER else 0
        if idx < len(PHE_CLASS_ORDER) - 1:
            larger = PHE_CLASS_ORDER[idx + 1]
            larger_A_plate = PHE_CLASSES[larger]["A_plate"]
            larger_max_A = PHE_CLASSES[larger]["max_area_single_frame"]
            alt = (f"Alternative: switch to a larger class ({larger}, "
                   f"{larger_A_plate:.2f} m²/plate, single-frame max "
                   f"≈ {larger_max_A:.0f} m²) for a more compact single-frame solution.")
        else:
            alt = (f"Even the largest standard class (M30, single-frame max "
                   f"≈ {max_A:.0f} m²) is exceeded — multi-frame is unavoidable.")
        warns.append(
            f"PHE area = {A:.0f} m² with {cls} plates ({Np} plates, "
            f"~{P['footprint_LWH_m'][0]:.1f} m frame) exceeds single-frame "
            f"limit of ~{max_pl} plates / {max_A:.0f} m². "
            f"Multi-frame installation required: ~{N_fr} parallel {cls} frames "
            f"of ~{max_pl} plates each. Total area unchanged "
            f"(set by Q / (U·LMTD)); multi-frame only packages it. "
            f"{alt}")
    elif Np > 0.85 * max_pl or A > 0.85 * max_A:
        # Approaching limit
        idx = PHE_CLASS_ORDER.index(cls) if cls in PHE_CLASS_ORDER else 0
        if idx < len(PHE_CLASS_ORDER) - 1:
            larger = PHE_CLASS_ORDER[idx + 1]
            warns.append(
                f"PHE design is near single-frame {cls} capacity "
                f"({Np}/{max_pl} plates, {A:.0f}/{max_A:.0f} m²). "
                f"Consider stepping up to {larger} for capacity headroom.")

    if U < 1000:
        warns.append(f"U = {U:.0f} W/(m²·K) is below the 1500-2500 W/(m²·K) "
                     f"typical for clean water-water GPHE; check fouling factors.")
    if L < 2.0:
        warns.append(f"LMTD = {L:.2f} K is very tight; design is at the thermal "
                     f"limit and small operating changes will shift performance.")
    if NTU > 8:
        if eps > 0.90:
            warns.append(
                f"NTU = {NTU:.2f} is high to achieve effectiveness ε = "
                f"{eps*100:.1f} % — this is INTENTIONAL for deep-cascade / "
                f"low-T DH (the required ε is fixed by the temperature "
                f"programme) but the design has little tolerance for fouling "
                f"or off-design operation. Verify manufacturer can guarantee "
                f"U = {U:.0f} W/(m²·K) with the specified fouling allowance.")
        else:
            warns.append(
                f"NTU = {NTU:.2f} is high (> 8) without commensurate "
                f"effectiveness gain (ε = {eps*100:.1f} %) — diminishing-"
                f"returns region; consider widening the approach.")
    if dpw > 60e3:
        warns.append(f"ΔP_water_HE = {dpw/1e3:.1f} kPa exceeds typical 50-60 kPa "
                     f"circulation-pump economic target.")
    if not R["ST"].get("feasible_12", True):
        warns.append(f"S&T alternative is infeasible at this approach: "
                     f"{R['ST'].get('F_note','')}")
    return warns


def is_infeasible(R):
    """Hard-failure check.  Returns (True, reason) if the design cannot be
    accepted under any conditions; otherwise (False, '')."""
    if R is None: return True, "calculation failed"
    P = R.get("PHE")
    if P is None: return True, "PHE sizing did not converge"
    if not math.isfinite(P["A_total_m2"]): return True, "PHE area is NaN/Inf"
    if not math.isfinite(P["LMTD_K"]):     return True, "LMTD is NaN/Inf"
    if P["LMTD_K"] < 0.5:    return True, f"LMTD = {P['LMTD_K']:.2f} K < 0.5 K"
    # Multi-frame is acceptable; only flag truly absurd designs as infeasible
    if P["N_frames_needed"] > 5:
        return True, (f"Even multi-frame M30 ({P['N_frames_needed']} frames) "
                      f"is impractical")
    if P["U_Wm2K"] < 400:    return True, f"U = {P['U_Wm2K']:.0f} W/(m²·K) < 400"
    return False, ""


# ==============================================================================
# PART 10 — SINGLE-CASE CALCULATION
# ==============================================================================
def run_case(option_label, T_DHS_out_input, approach_hot_in, approach_cold_in,
             plate_class_in="auto", verbose=True):
    """Compute one DH scenario.  Returns a results dict.  May raise ValueError
    on PHE crossover; caller is expected to catch and report."""
    # Resolve approaches (auto + override)
    a_hot, a_cold, app_notes = resolve_approaches(
        T_brine_in, T_DHS_out_input, approach_hot_in, approach_cold_in)

    T_HE_out = T_brine_in - a_hot
    rho_brine_in = rho_water(T_brine_in)
    m_brine = (Q_brine*1e-3)*rho_brine_in

    sumK, fittings_counts, fitting_K = K_fittings()

    # Initial guesses for fixed-point iteration
    m_water = m_brine
    T_HE_in = T_DHS_out_input - 0.5
    T_brine_out = T_HE_in + a_cold
    T_DHS_in = T_HE_out - 0.5

    for it in range(MAX_ITER):
        T_water_mean = 0.25*(T_HE_out + T_DHS_in + T_DHS_out_input + T_HE_in)
        water = water_props(T_water_mean)

        hyd_sup = pipe_hydraulics(m_water, water, D_inner, L_pipe, roughness, sumK)
        h_i_s = h_internal_Gnielinski(hyd_sup["Re"], water["Pr"], water["k"],
                                      D_inner, hyd_sup["f"])
        UA_sup, _ = UA_buried_pipe(L_pipe, D_inner, D_outer_steel,
                                   t_insulation, t_casing,
                                   k_steel, k_PUR, k_HDPE, k_soil,
                                   burial_depth, h_i_s)
        T_DHS_in_new = T_out_pipe(T_HE_out, T_soil_january, UA_sup,
                                  m_water, water["cp"])

        hyd_ret = pipe_hydraulics(m_water, water, D_inner, L_pipe, roughness, sumK)
        h_i_r = h_internal_Gnielinski(hyd_ret["Re"], water["Pr"], water["k"],
                                      D_inner, hyd_ret["f"])
        UA_ret, _ = UA_buried_pipe(L_pipe, D_inner, D_outer_steel,
                                   t_insulation, t_casing,
                                   k_steel, k_PUR, k_HDPE, k_soil,
                                   burial_depth, h_i_r)
        T_HE_in_new = T_out_pipe(T_DHS_out_input, T_soil_january, UA_ret,
                                 m_water, water["cp"])

        T_brine_out_new = T_HE_in_new + a_cold

        cp_w = water["cp"]
        cp_b = cp_water(0.5*(T_brine_in + T_brine_out_new))
        Q_HE = m_brine*cp_b*(T_brine_in - T_brine_out_new)
        # Guard against the divide-by-zero / negative denominator at low T_DHS_out
        denom = (T_HE_out - T_HE_in_new)*cp_w
        if denom <= 0:
            raise ValueError(f"Energy balance degenerate at T_DHS_out = "
                             f"{T_DHS_out_input:.1f} °C: T_HE_out-T_HE_in = "
                             f"{T_HE_out-T_HE_in_new:.2f} K")
        m_water_new = Q_HE / denom

        d_m   = abs(m_water_new - m_water)
        d_T_s = abs(T_DHS_in_new - T_DHS_in)
        d_T_r = abs(T_HE_in_new - T_HE_in)
        d_T_b = abs(T_brine_out_new - T_brine_out)

        m_water     = m_water_new
        T_DHS_in    = T_DHS_in_new
        T_HE_in     = T_HE_in_new
        T_brine_out = T_brine_out_new

        if d_m < ITER_TOL_KGS and d_T_s < 1e-3 and d_T_r < 1e-3 and d_T_b < 1e-3:
            break
    n_iter = it + 1

    # Final hydraulics
    T_sup_mean = 0.5*(T_HE_out + T_DHS_in)
    T_ret_mean = 0.5*(T_DHS_out_input + T_HE_in)
    sup_props = water_props(T_sup_mean)
    ret_props = water_props(T_ret_mean)
    hyd_sup = pipe_hydraulics(m_water, sup_props, D_inner, L_pipe, roughness, sumK)
    hyd_ret = pipe_hydraulics(m_water, ret_props, D_inner, L_pipe, roughness, sumK)

    h_i_s = h_internal_Gnielinski(hyd_sup["Re"], sup_props["Pr"], sup_props["k"],
                                  D_inner, hyd_sup["f"])
    h_i_r = h_internal_Gnielinski(hyd_ret["Re"], ret_props["Pr"], ret_props["k"],
                                  D_inner, hyd_ret["f"])
    UA_sup, _ = UA_buried_pipe(L_pipe, D_inner, D_outer_steel, t_insulation, t_casing,
                               k_steel, k_PUR, k_HDPE, k_soil, burial_depth, h_i_s)
    UA_ret, _ = UA_buried_pipe(L_pipe, D_inner, D_outer_steel, t_insulation, t_casing,
                               k_steel, k_PUR, k_HDPE, k_soil, burial_depth, h_i_r)
    Q_loss_sup = UA_sup*(T_sup_mean - T_soil_january)
    Q_loss_ret = UA_ret*(T_ret_mean - T_soil_january)

    # HE sizing — PHE (with class resolution) + S&T (for comparison)
    T_b_mean = 0.5*(T_brine_in + T_brine_out)
    T_w_mean = 0.5*(T_HE_in + T_HE_out)
    PHE_res, plate_notes = resolve_plate_class(
        plate_class_in,
        m_brine, water_props(T_b_mean), T_brine_in, T_brine_out,
        m_water, water_props(T_w_mean), T_HE_in, T_HE_out)
    # Append plate-class notes into the approach_notes list so console + Excel
    # both show them in one block.
    app_notes = list(app_notes) + plate_notes
    ST_res  = size_ST (m_brine, water_props(T_b_mean), T_brine_in, T_brine_out,
                       m_water, water_props(T_w_mean), T_HE_in, T_HE_out)

    # HE type recommendation
    HE_recommended, HE_reasons = recommend_HE_type(
        PHE_res, ST_res,
        T_max_C=T_brine_in, p_max_bar=10.0, fluid_type="clean water",
        duty_MW=PHE_res["Q_W"]/1e6)

    # Pump power
    HE_chosen = PHE_res if HE_recommended == "PHE" else ST_res
    dp_pump_Pa = hyd_sup["dp_total"] + hyd_ret["dp_total"] + HE_chosen["dp_water_Pa"]
    rho_pump = water_props(T_w_mean)["rho"]
    Q_vol = m_water/rho_pump
    H_m = dp_pump_Pa/(rho_pump*GRAV)
    P_hyd_W   = dp_pump_Pa*Q_vol
    P_shaft_W = P_hyd_W/eta_pump
    P_elec_W  = P_shaft_W/(eta_motor*eta_VFD)

    energy_kWh = P_elec_W/1000.0*FLH
    OPEX_excl = energy_kWh*ELEC_TARIFF_EUR_kWh_excl_VAT
    OPEX_incl = OPEX_excl*(1.0+VAT_RATE)

    return dict(
        option=option_label,
        T_brine_in=T_brine_in, T_brine_out=T_brine_out,
        T_HE_out=T_HE_out, T_HE_in=T_HE_in,
        T_DHS_in=T_DHS_in, T_DHS_out=T_DHS_out_input,
        Q_brine=Q_brine, m_brine=m_brine, m_water=m_water,
        approach_hot=a_hot, approach_cold=a_cold,
        approach_notes=app_notes,
        T_soil_january=T_soil_january,
        sumK=sumK, fittings_counts=fittings_counts, fitting_K=fitting_K,
        hyd_sup=hyd_sup, hyd_ret=hyd_ret,
        UA_sup=UA_sup, UA_ret=UA_ret,
        Q_loss_sup=Q_loss_sup, Q_loss_ret=Q_loss_ret,
        PHE=PHE_res, ST=ST_res,
        HE_recommended=HE_recommended, HE_reasons=HE_reasons,
        HE_chosen=HE_chosen,
        dp_pump_Pa=dp_pump_Pa, H_m=H_m, Q_vol=Q_vol,
        P_hyd_W=P_hyd_W, P_shaft_W=P_shaft_W, P_elec_W=P_elec_W,
        energy_kWh=energy_kWh, OPEX_excl=OPEX_excl, OPEX_incl=OPEX_incl,
        n_iter=n_iter,
    )


# ==============================================================================
# PART 11 — AUTO-FALLBACK WRAPPER
# ==============================================================================
def run_with_fallback(option_label, T_DHS_out_input,
                      approach_hot_in, approach_cold_in,
                      plate_class_in="auto", verbose=True):
    """Wrap run_case with auto-fallback for infeasible designs.

    Fallback strategy: if the auto-selected design is infeasible AND the user
    has not manually overridden the cold approach, the script LOOSENS the cold
    approach by 1 K and retries, up to a 6 K ceiling (loosening widens LMTD
    and reduces area).  All retries are logged.

    The fallback NEVER fires when approaches are manually set — the user's
    explicit choice is respected.
    """
    fallback_log = []
    # First attempt with whatever the user / auto picked
    try:
        R = run_case(option_label, T_DHS_out_input,
                     approach_hot_in, approach_cold_in,
                     plate_class_in=plate_class_in, verbose=False)
    except ValueError as e:
        # Hard math failure (crossover, etc.) — try fallback if we're on auto
        R = None
        fallback_log.append(f"Initial calculation failed: {e}")

    # Only attempt fallback if BOTH approaches are on auto (i.e., we have
    # freedom to adjust without overriding user intent).
    auto_mode = (approach_hot_in is None) and (approach_cold_in is None)

    if R is not None:
        infeasible, reason = is_infeasible(R)
    else:
        infeasible, reason = True, fallback_log[-1] if fallback_log else "unknown"

    fallback_used = False
    cold_used = None
    if infeasible and auto_mode:
        # Determine the auto-picked cold approach as the starting point
        _, auto_cold0, _, _ = auto_approaches(T_brine_in, T_DHS_out_input)
        cold_try = auto_cold0
        fallback_log.append(
            f"Initial design infeasible ({reason}). Auto-fallback engaged: "
            f"loosening cold approach from auto-default {auto_cold0:.1f} K "
            f"in 1 K steps up to 6 K ceiling.")
        while infeasible and cold_try < 6.0:
            cold_try += 1.0
            try:
                R_try = run_case(option_label, T_DHS_out_input,
                                 approach_hot_in, cold_try,
                                 plate_class_in=plate_class_in, verbose=False)
                inf2, reason2 = is_infeasible(R_try)
                fallback_log.append(
                    f"  retry with cold approach = {cold_try:.1f} K  →  "
                    f"PHE A = {R_try['PHE']['A_total_m2']:.0f} m², "
                    f"U = {R_try['PHE']['U_Wm2K']:.0f}, "
                    f"LMTD = {R_try['PHE']['LMTD_K']:.2f} K, "
                    f"NTU = {R_try['PHE']['NTU']:.2f}"
                    + ("  → ACCEPTED" if not inf2 else f"  → still infeasible ({reason2})"))
                if not inf2:
                    R = R_try
                    infeasible = False
                    cold_used = cold_try
                    fallback_used = True
                    break
                else:
                    R = R_try  # keep the latest even if still flagged
            except ValueError as e:
                fallback_log.append(f"  retry with cold approach = "
                                    f"{cold_try:.1f} K failed: {e}")
        if infeasible and R is not None:
            fallback_log.append(
                f"  Fallback exhausted at cold approach = {cold_try:.1f} K; "
                f"final design still flagged. Reporting as-is with warnings.")
    elif infeasible and not auto_mode:
        fallback_log.append(
            f"Design flagged as infeasible ({reason}) but approaches are "
            f"MANUALLY OVERRIDDEN; fallback skipped per user intent. Reporting "
            f"as-is with warnings.")

    # Generate sanity warnings regardless of fallback status
    warnings = design_sanity_warnings(R) if R is not None else []

    if R is not None:
        # If the fallback engaged (regardless of success), the approach_notes
        # currently say "MANUALLY SET" — that is misleading because the loosening
        # was done automatically.  Overwrite with the truthful description.
        if fallback_log and auto_mode:
            _, auto_cold0, _, _ = auto_approaches(T_brine_in, T_DHS_out_input)
            R["approach_notes"] = [
                n for n in R["approach_notes"]
                if not n.startswith("COLD APPROACH")
            ]
            if fallback_used:
                R["approach_notes"].append(
                    f"COLD APPROACH: {R['approach_cold']:.1f} K "
                    f"(LOOSENED BY AUTO-FALLBACK from auto-default "
                    f"{auto_cold0:.1f} K because initial design was infeasible)")
            else:
                R["approach_notes"].append(
                    f"COLD APPROACH: {R['approach_cold']:.1f} K "
                    f"(auto-fallback exhausted: auto-default was "
                    f"{auto_cold0:.1f} K; reporting last attempt with warnings)")
        R["fallback_log"] = fallback_log
        R["fallback_used"] = fallback_used
        R["warnings"] = warnings
    return R


# ==============================================================================
# PART 12 — PER-OPTION CONSOLE OUTPUT
# ==============================================================================
def print_case(R):
    if R is None:
        print("  (case failed — no results)")
        return
    print("="*78)
    print(f" OPTION {R['option']}  —  T_DHS_out = {R['T_DHS_out']:.1f} °C")
    print("="*78)
    print(" Approach selection:")
    for n in R["approach_notes"]:
        print(f"   {n}")
    if R.get("fallback_log"):
        print(" Auto-fallback log:")
        for n in R["fallback_log"]:
            print(f"   {n}")
    print(f" Iteration converged in {R['n_iter']} passes")
    print()
    print(" Working-fluid temperatures:")
    print(f"   T_HE_out     = {R['T_HE_out']:7.3f} °C   (= T_brine_in − {R['approach_hot']:.1f} K)")
    print(f"   T_DHS_in     = {R['T_DHS_in']:7.3f} °C   (ΔT_supply = {R['T_HE_out']-R['T_DHS_in']:.3f} K)")
    print(f"   T_DHS_out    = {R['T_DHS_out']:7.3f} °C   (FIXED by user)")
    print(f"   T_HE_in      = {R['T_HE_in']:7.3f} °C   (ΔT_return = {R['T_DHS_out']-R['T_HE_in']:.3f} K)")
    print(f"   T_brine_out  = {R['T_brine_out']:7.3f} °C   (= T_HE_in + {R['approach_cold']:.1f} K)")
    print(f"   ṁ_brine = {R['m_brine']:6.2f} kg/s   ṁ_water = {R['m_water']:6.2f} kg/s  ({R['Q_vol']*1000:.2f} L/s)")
    print(f"   Q_HE    = {R['PHE']['Q_W']/1e6:.3f} MW")
    print()
    print(" Pipeline:")
    for tag, h, q in [("Supply", R["hyd_sup"], R["Q_loss_sup"]),
                      ("Return", R["hyd_ret"], R["Q_loss_ret"])]:
        print(f"   {tag}: v={h['v']:.3f} m/s, Re={h['Re']:.2e}, "
              f"ΔP_total={h['dp_total']/1e5:.4f} bar, heat loss = {q/1000:.2f} kW")
    print()
    print(" HE sizing (both calculated):")
    P = R["PHE"]
    fits_tag = "" if P["fits_single_frame"] else f"  → multi-frame: {P['N_frames_needed']} parallel frames"
    print(f"   PHE (gasketed, {P['plate_class']} plates "
          f"{P['plate_W_m']:.2f}×{P['plate_H_m']:.2f} m, {P['A_plate_m2']:.2f} m²/plate):")
    print(f"      U = {P['U_Wm2K']:.0f} W/(m²·K), "
          f"A = {P['A_total_m2']:.1f} m², "
          f"{P['N_plates']} plates, footprint ≈ "
          f"{P['footprint_LWH_m'][0]:.1f}×{P['footprint_LWH_m'][1]:.2f}×"
          f"{P['footprint_LWH_m'][2]:.2f} m, "
          f"weight ≈ {P['weight_kg']:.0f} kg{fits_tag}")
    print(f"      LMTD = {P['LMTD_K']:.2f} K, NTU = {P['NTU']:.2f}, "
          f"ε = {P['effectiveness']*100:.1f} %, "
          f"ΔP water = {P['dp_water_Pa']/1e3:.1f} kPa")
    r = R["ST"]
    feas = r.get("feasible_12", True)
    tag = "" if feas else "  [INFEASIBLE — F=1 idealisation]"
    print(f"   {r['type']}: U = {r['U_Wm2K']:.0f} W/(m²·K), "
          f"A = {r['A_total_m2']:.1f} m², "
          f"footprint ≈ {r['footprint_LWH_m'][0]:.1f}×"
          f"{r['footprint_LWH_m'][1]:.2f}×"
          f"{r['footprint_LWH_m'][2]:.2f} m, "
          f"weight ≈ {r['weight_kg']:.0f} kg{tag}")
    print(f"      LMTD = {r['LMTD_K']:.2f} K, NTU = {r['NTU']:.2f}, "
          f"ε = {r['effectiveness']*100:.1f} %, "
          f"ΔP water = {r['dp_water_Pa']/1e3:.1f} kPa")
    if not feas: print(f"      {r['F_note']}")
    print()
    print(f" Recommendation: {R['HE_recommended']}")
    for j, reason in enumerate(R["HE_reasons"], 1):
        print(f"   ({j}) {reason}")
    print()
    print(" Pump & OPEX:")
    print(f"   Pump ΔP = {R['dp_pump_Pa']/1e5:.3f} bar   H = {R['H_m']:.2f} m")
    print(f"   P_elec  = {R['P_elec_W']/1000:.2f} kW   "
          f"annual energy = {R['energy_kWh']:,.0f} kWh/yr")
    print(f"   OPEX excl. VAT = {R['OPEX_excl']:>10,.2f} EUR/yr")
    print(f"   OPEX incl. VAT = {R['OPEX_incl']:>10,.2f} EUR/yr")
    if R.get("warnings"):
        print()
        print(" SANITY WARNINGS:")
        for w in R["warnings"]:
            print(f"   ⚠  {w}")
    print("="*78)
    print()


# ==============================================================================
# PART 13 — CROSS-OPTION COMPARISON TABLE
# ==============================================================================
def print_comparison(results):
    """Print a side-by-side comparison of all enabled options."""
    enabled = [r for r in results if r is not None]
    if not enabled: return
    print()
    print("#"*78)
    print(" CROSS-OPTION COMPARISON")
    print("#"*78)
    rows = [
        ("T_DHS_out          [°C]",   "{:.2f}",     "T_DHS_out"),
        ("Hot approach       [K]",    "{:.1f}",     "approach_hot"),
        ("Cold approach      [K]",    "{:.1f}",     "approach_cold"),
        ("PHE plate class    [—]",    "{}",         lambda r: r["PHE"]["plate_class"]),
        ("Plate dims    [m × m]",     "{}",         lambda r: f"{r['PHE']['plate_W_m']:.2f}×{r['PHE']['plate_H_m']:.2f}"),
        ("Fits single frame  [—]",    "{}",         lambda r: "yes" if r["PHE"]["fits_single_frame"] else f"no ({r['PHE']['N_frames_needed']} frames)"),
        ("T_HE_out           [°C]",   "{:.2f}",     "T_HE_out"),
        ("T_HE_in            [°C]",   "{:.2f}",     "T_HE_in"),
        ("T_brine_out        [°C]",   "{:.2f}",     "T_brine_out"),
        ("ṁ_water           [kg/s]",  "{:.2f}",     "m_water"),
        ("Q_HE               [MW]",   "{:.3f}",     lambda r: r["PHE"]["Q_W"]/1e6),
        ("LMTD               [K]",    "{:.2f}",     lambda r: r["PHE"]["LMTD_K"]),
        ("PHE U          [W/(m²·K)]", "{:.0f}",     lambda r: r["PHE"]["U_Wm2K"]),
        ("PHE area           [m²]",   "{:.1f}",     lambda r: r["PHE"]["A_total_m2"]),
        ("PHE plates         [—]",    "{:d}",       lambda r: r["PHE"]["N_plates"]),
        ("PHE frame length   [m]",    "{:.2f}",     lambda r: r["PHE"]["footprint_LWH_m"][0]),
        ("Effectiveness ε    [%]",    "{:.1f}",     lambda r: r["PHE"]["effectiveness"]*100),
        ("NTU                [—]",    "{:.2f}",     lambda r: r["PHE"]["NTU"]),
        ("ΔP_water_HE       [kPa]",   "{:.1f}",     lambda r: r["PHE"]["dp_water_Pa"]/1e3),
        ("Pump P_elec        [kW]",   "{:.2f}",     lambda r: r["P_elec_W"]/1000),
        ("Annual energy    [kWh/yr]", "{:,.0f}",    "energy_kWh"),
        ("OPEX excl. VAT  [EUR/yr]",  "{:,.2f}",    "OPEX_excl"),
        ("OPEX incl. VAT  [EUR/yr]",  "{:,.2f}",    "OPEX_incl"),
    ]
    hdr = " {:<30}" + "".join("  Option {:<6}".format(r["option"]) for r in enabled)
    print(hdr.format(""))
    print(" "+"-"*78)
    for label, fmt, key in rows:
        vals = []
        for r in enabled:
            v = key(r) if callable(key) else r[key]
            try: vals.append(fmt.format(v))
            except (ValueError, TypeError): vals.append(str(v))
        print(" {:<30}".format(label) + "".join("  {:<13}".format(v) for v in vals))
    print("#"*78)
    print()


# ==============================================================================
# PART 14 — EXCEL OUTPUT  (one sheet per option + sensitivity summary)
# ==============================================================================
def write_excel(results, save_dir):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    wb = Workbook()
    # remove default sheet; we'll add ours
    wb.remove(wb.active)

    hdr_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    sec_font = Font(name="Arial", size=12, bold=True)
    unit_font = Font(name="Arial", size=10, italic=True, color="555555")
    norm_font = Font(name="Arial", size=11)
    warn_font = Font(name="Arial", size=10, italic=True, color="C00000")
    hdr_fill = PatternFill("solid", fgColor="305496")
    in_fill  = PatternFill("solid", fgColor="FFF2CC")
    out_fill = PatternFill("solid", fgColor="DDEBF7")
    h_fill   = PatternFill("solid", fgColor="FCE4D6")
    p_fill   = PatternFill("solid", fgColor="E2EFDA")
    s_fill   = PatternFill("solid", fgColor="F2D7B6")
    e_fill   = PatternFill("solid", fgColor="FFE699")
    n_fill   = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def wr(ws, r, no, name, val, unit, note, fill):
        ws.cell(r,1,no).font = norm_font
        ws.cell(r,2,name).font = norm_font
        c = ws.cell(r,3,val); c.font = norm_font
        if isinstance(val, float):
            c.number_format = "0.0000" if abs(val)<1 else ("0.000" if abs(val)<1000 else "0.0")
        ws.cell(r,4,unit).font = unit_font
        ws.cell(r,5,note).font = unit_font
        for cc in range(1,6):
            ws.cell(r,cc).fill = fill; ws.cell(r,cc).border = border

    # ----- One sheet per option -----
    for R in results:
        if R is None: continue
        ws = wb.create_sheet(title=f"Option {R['option']}")
        for col,w in enumerate([6, 46, 18, 12, 50], 1):
            ws.column_dimensions[chr(64+col)].width = w
        ws["A1"] = f"Geothermal DH — Option {R['option']}  (T_DHS_out = {R['T_DHS_out']:.1f} °C)"
        ws["A1"].font = Font(name="Arial", size=14, bold=True); ws.merge_cells("A1:E1")

        r = 3
        # Approach notes
        ws.cell(r,1,"APPROACH SELECTION").font = sec_font; r += 1
        for n in R["approach_notes"]:
            ws.cell(r,2,n).font = norm_font
            for cc in range(1,6): ws.cell(r,cc).fill = n_fill; ws.cell(r,cc).border = border
            r += 1
        if R.get("fallback_log"):
            r += 1
            ws.cell(r,1,"AUTO-FALLBACK LOG").font = sec_font; r += 1
            for n in R["fallback_log"]:
                ws.cell(r,2,n).font = norm_font
                for cc in range(1,6): ws.cell(r,cc).fill = n_fill; ws.cell(r,cc).border = border
                r += 1
        r += 1

        # Inputs
        ws.cell(r,1,"INPUTS").font = sec_font; r += 1
        for c,h in enumerate(["#","Parameter","Value","Unit","Note / source"],1):
            ws.cell(r,c,h).font = hdr_font; ws.cell(r,c).fill = hdr_fill; ws.cell(r,c).border = border
        r += 1
        inputs_rows = [
            ("1","T_brine_in",       T_brine_in,    "°C",    "wellhead"),
            ("2","Q_brine",          Q_brine,       "L/s",   "wellhead"),
            ("3","T_DHS_out",        R["T_DHS_out"],"°C",    "FIXED for this option"),
            ("4","approach_hot",     R["approach_hot"],"K",  "auto or manual override"),
            ("5","approach_cold",    R["approach_cold"],"K", "auto or manual override"),
            ("6","T_soil_january",   T_soil_january,"°C",    "local met. service, 1 m depth, January"),
            ("7","L_pipe",           L_pipe,        "m",     "one way"),
            ("8","burial_depth",     burial_depth,  "m",     ""),
            ("9","D_inner",          D_inner,       "m",     "DN 200 EN 253"),
            ("10","t_insulation",    t_insulation,  "m",     "Logstor series 2 DN200/355"),
            ("11","FLH",             FLH,           "h/yr",  "common to all options"),
            ("12","Tariff excl. VAT",ELEC_TARIFF_EUR_kWh_excl_VAT,"EUR/kWh",
                  "Eurostat nrg_pc_205 Croatia H1 2025"),
            ("13","VAT rate",        VAT_RATE,      "—",     "Croatian 25 %"),
        ]
        for i in inputs_rows: wr(ws, r, *i, in_fill); r += 1
        r += 1

        # Working-fluid temperatures
        ws.cell(r,1,"WORKING-FLUID TEMPERATURES").font = sec_font; r += 1
        for c,h in enumerate(["#","Quantity","Value","Unit","Note"],1):
            ws.cell(r,c,h).font = hdr_font; ws.cell(r,c).fill = hdr_fill; ws.cell(r,c).border = border
        r += 1
        for row in [
            ("T1","T_HE_out (HE → supply pipe)",  R["T_HE_out"], "°C", f"= T_brine_in − {R['approach_hot']:.1f} K"),
            ("T2","T_DHS_in (supply pipe → DHS)", R["T_DHS_in"], "°C", "computed"),
            ("T3","ΔT supply pipe", R["T_HE_out"]-R["T_DHS_in"], "K", ""),
            ("T4","T_DHS_out (DHS → return pipe)",R["T_DHS_out"],"°C", "FIXED"),
            ("T5","T_HE_in (return pipe → HE)",   R["T_HE_in"],  "°C", "computed"),
            ("T6","ΔT return pipe", R["T_DHS_out"]-R["T_HE_in"], "K", ""),
            ("T7","T_brine_out (HE → injection)", R["T_brine_out"],"°C",f"= T_HE_in + {R['approach_cold']:.1f} K"),
            ("T8","ṁ_brine",        R["m_brine"],  "kg/s",       ""),
            ("T9","ṁ_water",        R["m_water"],  "kg/s",       "from energy balance"),
        ]:
            wr(ws, r, *row, out_fill); r += 1
        r += 1

        # Hydraulics
        ws.cell(r,1,"HYDRAULICS").font = sec_font; r += 1
        for c,h in enumerate(["#","Quantity","Supply","Unit","Return"],1):
            ws.cell(r,c,h).font = hdr_font; ws.cell(r,c).fill = hdr_fill; ws.cell(r,c).border = border
        r += 1
        for tag,name,vs,u,vr in [
            ("H1","v",            R["hyd_sup"]["v"],          "m/s", R["hyd_ret"]["v"]),
            ("H2","Re",           R["hyd_sup"]["Re"],         "—",   R["hyd_ret"]["Re"]),
            ("H3","f (Colebrook)",R["hyd_sup"]["f"], "—",            R["hyd_ret"]["f"]),
            ("H4","ΔP friction",  R["hyd_sup"]["dp_friction"]/1e5,"bar", R["hyd_ret"]["dp_friction"]/1e5),
            ("H5","ΔP minor",     R["hyd_sup"]["dp_minor"]/1e5,   "bar", R["hyd_ret"]["dp_minor"]/1e5),
            ("H6","ΔP total",     R["hyd_sup"]["dp_total"]/1e5,   "bar", R["hyd_ret"]["dp_total"]/1e5),
            ("H7","UA buried pipe",R["UA_sup"],                   "W/K", R["UA_ret"]),
            ("H8","Heat loss",    R["Q_loss_sup"]/1000,           "kW",  R["Q_loss_ret"]/1000),
        ]:
            wr(ws, r, tag, name, vs, u, f"return: {vr:.4g}", h_fill); r += 1
        r += 1

        # PHE
        ws.cell(r,1,"PLATE HEAT EXCHANGER").font = sec_font; r += 1
        for c,h in enumerate(["#","Quantity","Value","Unit","Note"],1):
            ws.cell(r,c,h).font = hdr_font; ws.cell(r,c).fill = hdr_fill; ws.cell(r,c).border = border
        r += 1
        P = R["PHE"]
        for row in [
            ("E0a","Plate class",     P["plate_class"],       "—","Alfa Laval Industrial Line"),
            ("E0b","Plate dimensions",
                  f'{P["plate_W_m"]:.3f} × {P["plate_H_m"]:.3f}', "m × m",
                  f'A_plate = {P["A_plate_m2"]:.3f} m²'),
            ("E0c","Single-frame fit",
                  "yes" if P["fits_single_frame"] else f'no, requires {P["N_frames_needed"]} parallel frames',
                  "—",
                  f'limit: {P["max_plates_per_frame"]} plates / {P["max_area_single_frame_m2"]:.0f} m² per frame'),
            ("E1","Duty Q",          P["Q_W"]/1e6,           "MW",""),
            ("E2","Approach hot/cold",
                  f'{P["approach_hot"]:.2f} / {P["approach_cold"]:.2f}', "K",""),
            ("E3","LMTD",            P["LMTD_K"],            "K","counter-current"),
            ("E4","U overall",       P["U_Wm2K"],            "W/(m²·K)","Martin (1996)"),
            ("E5","Area required",   P["A_req_m2"],          "m²",""),
            ("E6","Number of plates",P["N_plates"],          "—",""),
            ("E7","Installed area",  P["A_total_m2"],        "m²",""),
            ("E8","NTU",             P["NTU"],               "—",""),
            ("E9","Effectiveness ε", P["effectiveness"]*100, "%",""),
            ("E10","ΔP brine side",  P["dp_brine_Pa"]/1e3,   "kPa",""),
            ("E11","ΔP water side",  P["dp_water_Pa"]/1e3,   "kPa",""),
            ("E12","Footprint L×W×H",
                  f'{P["footprint_LWH_m"][0]:.2f}×{P["footprint_LWH_m"][1]:.2f}×{P["footprint_LWH_m"][2]:.2f}',
                  "m","Alfa Laval Industrial line"),
            ("E13","Weight",         P["weight_kg"], "kg",""),
        ]:
            wr(ws, r, *row, p_fill); r += 1
        r += 1

        # S&T
        ws.cell(r,1,"SHELL-AND-TUBE (calculated, TEMA E, 1-2)").font = sec_font; r += 1
        for c,h in enumerate(["#","Quantity","Value","Unit","Note"],1):
            ws.cell(r,c,h).font = hdr_font; ws.cell(r,c).fill = hdr_fill; ws.cell(r,c).border = border
        r += 1
        S = R["ST"]
        for row in [
            ("S1","Duty Q",          S["Q_W"]/1e6,           "MW",""),
            ("S2","LMTD",            S["LMTD_K"],            "K",""),
            ("S3","F correction",    S["F_correction"],      "—","1-2 (Bowman 1940)"),
            ("S4","U overall",       S["U_Wm2K"],            "W/(m²·K)","Kern (1950)"),
            ("S5","Area required",   S["A_req_m2"],          "m²",""),
            ("S6","Installed area",  S["A_total_m2"],        "m²",""),
            ("S7","Number of tubes", S["N_tubes"],           "—","19.05×1.65 mm"),
            ("S8","Tube length",     S["L_tube_m"],          "m","may exceed std 4.88 m"),
            ("S9","Shell ID",        S["D_shell_m"],         "m",""),
            ("S10","Effectiveness ε",S["effectiveness"]*100, "%",""),
            ("S11","Feasibility 1-2 TEMA E", "yes" if S.get("feasible_12",True) else "NO", "",
              S.get("F_note","")),
        ]:
            wr(ws, r, *row, s_fill); r += 1
        r += 1

        # Recommendation
        ws.cell(r,1,f"HE RECOMMENDATION: {R['HE_recommended']}").font = sec_font; r += 1
        for j, reason in enumerate(R["HE_reasons"], 1):
            wr(ws, r, f"R{j}", reason,"","","", out_fill); r += 1
        r += 1

        # Pump & OPEX
        ws.cell(r,1,"PUMP & OPEX").font = sec_font; r += 1
        for c,h in enumerate(["#","Quantity","Value","Unit","Note"],1):
            ws.cell(r,c,h).font = hdr_font; ws.cell(r,c).fill = hdr_fill; ws.cell(r,c).border = border
        r += 1
        for row in [
            ("P1","Pump ΔP total",   R["dp_pump_Pa"]/1e5,    "bar","supply + return + HE"),
            ("P2","Head H",          R["H_m"],               "m",""),
            ("P3","Volumetric flow", R["Q_vol"]*3600,        "m³/h",""),
            ("P4","P_elec",          R["P_elec_W"]/1000,     "kW",""),
            ("P5","FLH",             FLH,                    "h/yr",""),
            ("P6","Annual energy",   R["energy_kWh"],        "kWh/yr",""),
            ("P7","OPEX excl. VAT",  R["OPEX_excl"],         "EUR/yr",""),
            ("P8","OPEX incl. VAT",  R["OPEX_incl"],         "EUR/yr",""),
        ]:
            wr(ws, r, *row, e_fill); r += 1
        r += 1

        # Sanity warnings
        if R.get("warnings"):
            ws.cell(r,1,"SANITY WARNINGS").font = sec_font; r += 1
            for w in R["warnings"]:
                ws.cell(r,2,"⚠  "+w).font = warn_font
                for cc in range(1,6): ws.cell(r,cc).border = border
                r += 1

    # ----- Sensitivity summary sheet -----
    ws = wb.create_sheet(title="Sensitivity summary")
    for col,w in enumerate([28, 14, 14, 14], 1):
        ws.column_dimensions[chr(64+col)].width = w
    ws["A1"] = "Cross-option comparison summary"
    ws["A1"].font = Font(name="Arial", size=14, bold=True); ws.merge_cells("A1:D1")
    enabled = [r for r in results if r is not None]

    # Header row
    r = 3
    ws.cell(r,1,"Quantity").font = hdr_font; ws.cell(r,1).fill = hdr_fill
    ws.cell(r,1).border = border
    for j, R in enumerate(enabled, 2):
        c = ws.cell(r, j, f"Option {R['option']}")
        c.font = hdr_font; c.fill = hdr_fill; c.border = border
        c.alignment = Alignment(horizontal="center")
    r += 1

    rows = [
        ("T_DHS_out [°C]",        lambda r: r["T_DHS_out"],                "{:.2f}"),
        ("Hot approach [K]",      lambda r: r["approach_hot"],             "{:.1f}"),
        ("Cold approach [K]",     lambda r: r["approach_cold"],            "{:.1f}"),
        ("PHE plate class",       lambda r: r["PHE"]["plate_class"],       "{}"),
        ("Plate dims [m × m]",    lambda r: f"{r['PHE']['plate_W_m']:.2f}×{r['PHE']['plate_H_m']:.2f}", "{}"),
        ("Fits single frame",     lambda r: "yes" if r["PHE"]["fits_single_frame"] else f"no ({r['PHE']['N_frames_needed']} frames)", "{}"),
        ("T_HE_out [°C]",         lambda r: r["T_HE_out"],                 "{:.2f}"),
        ("T_HE_in [°C]",          lambda r: r["T_HE_in"],                  "{:.2f}"),
        ("T_brine_out [°C]",      lambda r: r["T_brine_out"],              "{:.2f}"),
        ("ṁ_water [kg/s]",        lambda r: r["m_water"],                  "{:.2f}"),
        ("Q_HE [MW]",             lambda r: r["PHE"]["Q_W"]/1e6,           "{:.3f}"),
        ("LMTD [K]",              lambda r: r["PHE"]["LMTD_K"],            "{:.2f}"),
        ("PHE U [W/(m²·K)]",      lambda r: r["PHE"]["U_Wm2K"],            "{:.0f}"),
        ("PHE area [m²]",         lambda r: r["PHE"]["A_total_m2"],        "{:.1f}"),
        ("PHE plates [—]",        lambda r: r["PHE"]["N_plates"],          "{:d}"),
        ("PHE frame length [m]",  lambda r: r["PHE"]["footprint_LWH_m"][0],"{:.2f}"),
        ("Effectiveness ε [%]",   lambda r: r["PHE"]["effectiveness"]*100, "{:.1f}"),
        ("NTU [—]",               lambda r: r["PHE"]["NTU"],               "{:.2f}"),
        ("ΔP_water_HE [kPa]",     lambda r: r["PHE"]["dp_water_Pa"]/1e3,   "{:.1f}"),
        ("Pump P_elec [kW]",      lambda r: r["P_elec_W"]/1000,            "{:.2f}"),
        ("Annual energy [kWh/yr]",lambda r: r["energy_kWh"],               "{:,.0f}"),
        ("OPEX excl. VAT [EUR/yr]",lambda r: r["OPEX_excl"],               "{:,.2f}"),
        ("OPEX incl. VAT [EUR/yr]",lambda r: r["OPEX_incl"],               "{:,.2f}"),
    ]
    for label, getter, fmt in rows:
        ws.cell(r,1,label).font = norm_font
        ws.cell(r,1).fill = n_fill; ws.cell(r,1).border = border
        for j, R in enumerate(enabled, 2):
            val = getter(R)
            try: txt = fmt.format(val)
            except (ValueError, TypeError): txt = str(val)
            c = ws.cell(r, j, txt); c.font = norm_font
            c.border = border
            c.alignment = Alignment(horizontal="right")
        r += 1

    path = os.path.join(save_dir, "geothermal_HE_sa_results.xlsx")
    wb.save(path); return path


# ==============================================================================
# PART 15 — SCHEMATIC DRAWERS (parameterized by option label)
# ==============================================================================
def draw_PHE_scheme(R, save_dir):
    import matplotlib.pyplot as plt
    import matplotlib.patches as mp
    P = R["PHE"]
    fig, ax = plt.subplots(figsize=(12, 6.0))
    ax.set_xlim(0, 14); ax.set_ylim(0, 7); ax.axis("off")
    ax.set_title(f"Plate Heat Exchanger — Option {R['option']} "
                 f"(T_DHS_out = {R['T_DHS_out']:.1f} °C)",
                 fontsize=13, fontweight="bold", pad=14)
    phe_x, phe_y, phe_w, phe_h = 5.6, 1.2, 2.8, 4.5
    ax.add_patch(mp.Rectangle((phe_x,phe_y),phe_w,phe_h,fc="white",ec="black",lw=2.0))
    band_w = 0.55
    ax.add_patch(mp.Rectangle((phe_x+0.20, phe_y+0.20), band_w, phe_h-0.40,
                              fc="#FBE5E5", ec="#C0504D", lw=0.8))
    ax.add_patch(mp.Rectangle((phe_x+phe_w-0.20-band_w, phe_y+0.20), band_w, phe_h-0.40,
                              fc="#E1ECF7", ec="#1F4E79", lw=0.8))
    for i in range(11):
        ystr = phe_y + 0.30 + i*(phe_h-0.60)/10
        ax.plot([phe_x+0.20+band_w+0.05, phe_x+phe_w-0.20-band_w-0.05],
                [ystr,ystr], "-", color="#888888", lw=0.5)
    y_top = phe_y+phe_h-0.70; y_bot = phe_y+0.70
    ax.annotate("", xy=(phe_x, y_top), xytext=(1.6, y_top),
                arrowprops=dict(arrowstyle="-|>,head_width=0.35,head_length=0.4",
                                color="#A52A2A", lw=4.5))
    ax.annotate("", xy=(1.6, y_bot), xytext=(phe_x, y_bot),
                arrowprops=dict(arrowstyle="-|>,head_width=0.35,head_length=0.4",
                                color="#5B1A1A", lw=4.5))
    ax.text(1.6, y_top+0.45, f"{R['T_brine_in']:.2f} °C", ha="center",
            fontsize=14, fontweight="bold", color="#A52A2A")
    ax.text(1.6, y_bot-0.50, f"{R['T_brine_out']:.2f} °C", ha="center",
            fontsize=14, fontweight="bold", color="#5B1A1A")
    cx_l = phe_x-1.9
    ax.text(cx_l, phe_y+phe_h-0.20, "Geothermal brine\n(HE primary)", ha="center",
            fontsize=10.5, fontweight="bold", color="#A52A2A")
    ax.text(cx_l, phe_y+phe_h/2+0.55,
            f"ṁ = {R['m_brine']*3.6:.2f} t/h\n   = {R['m_brine']:.2f} kg/s",
            ha="center", fontsize=10.5)
    ax.text(cx_l, phe_y+phe_h/2-0.10,
            f"ΔP$_{{HE}}$ = {P['dp_brine_Pa']/1000:.2f} kPa",
            ha="center", fontsize=10.5)
    ax.text(cx_l, phe_y+phe_h/2-0.70,
            f"Q = {P['Q_W']/1e6:.2f} MW", ha="center", fontsize=11, fontweight="bold")
    ax.annotate("", xy=(12.4, y_top), xytext=(phe_x+phe_w, y_top),
                arrowprops=dict(arrowstyle="-|>,head_width=0.35,head_length=0.4",
                                color="#D88080", lw=4.5))
    ax.annotate("", xy=(phe_x+phe_w, y_bot), xytext=(12.4, y_bot),
                arrowprops=dict(arrowstyle="-|>,head_width=0.35,head_length=0.4",
                                color="#1F4E79", lw=4.5))
    ax.text(12.4, y_top+0.45, f"{R['T_HE_out']:.2f} °C", ha="center",
            fontsize=14, fontweight="bold", color="#C0504D")
    ax.text(12.4, y_bot-0.50, f"{R['T_HE_in']:.2f} °C", ha="center",
            fontsize=14, fontweight="bold", color="#1F4E79")
    ax.text(12.4, y_top-0.40, "→ to DHS", ha="center", fontsize=9, style="italic", color="#555555")
    ax.text(12.4, y_bot+0.30, "← from DHS", ha="center", fontsize=9, style="italic", color="#555555")
    cx_r = phe_x+phe_w+1.9
    ax.text(cx_r, phe_y+phe_h-0.20, "Working fluid (water)\n(HE secondary)", ha="center",
            fontsize=10.5, fontweight="bold", color="#1F4E79")
    ax.text(cx_r, phe_y+phe_h/2+0.55,
            f"ṁ = {R['m_water']*3.6:.2f} t/h\n   = {R['m_water']:.2f} kg/s",
            ha="center", fontsize=10.5)
    ax.text(cx_r, phe_y+phe_h/2-0.10,
            f"ΔP$_{{HE}}$ = {P['dp_water_Pa']/1000:.2f} kPa",
            ha="center", fontsize=10.5)
    ax.text(cx_r, phe_y+phe_h/2-0.70,
            f"Q = {P['Q_W']/1e6:.2f} MW", ha="center", fontsize=11, fontweight="bold")
    foot = (f"PHE (gasketed, 316L SS, chevron {PHE_chevron_deg:.0f}°, "
            f"{P['plate_class']} plates {P['plate_W_m']:.2f}×{P['plate_H_m']:.2f} m):  "
            f"A = {P['A_total_m2']:.1f} m² ({P['N_plates']} plates)   |   "
            f"U = {P['U_Wm2K']:.0f} W/(m²·K)   |   "
            f"LMTD = {P['LMTD_K']:.2f} K   |   "
            f"NTU = {P['NTU']:.2f}   |   "
            f"ε = {P['effectiveness']*100:.1f} %   |   "
            f"approach = {R['approach_hot']:.1f}/{R['approach_cold']:.1f} K"
            + ("" if P["fits_single_frame"] else
               f"   |   multi-frame: {P['N_frames_needed']} × {P['plate_class']} parallel"))
    ax.text(7, 0.45, foot, ha="center", fontsize=9.5,
            bbox=dict(boxstyle="round,pad=0.4", fc="#F2F2F2", ec="black"))
    plt.tight_layout()
    path = os.path.join(save_dir, f"PHE_scheme_{R['option']}.png")
    plt.savefig(path, dpi=200, bbox_inches="tight"); plt.close(fig)
    return path


def draw_system_scheme(R, save_dir):
    import matplotlib.pyplot as plt
    import matplotlib.patches as mp
    fig, ax = plt.subplots(figsize=(15, 7.0))
    ax.set_xlim(0, 22); ax.set_ylim(0, 9); ax.axis("off")
    ax.set_title(f"System schematic — Option {R['option']} "
                 f"(T_DHS_out = {R['T_DHS_out']:.1f} °C)",
                 fontsize=14, fontweight="bold", pad=14)
    P = R["PHE"]

    # HE box
    he_x, he_y, he_w, he_h = 4.0, 1.8, 1.7, 5.4
    ax.add_patch(mp.Rectangle((he_x, he_y), he_w, he_h,
                              fc="white", ec="black", lw=1.8))
    band_w = 0.32
    ax.add_patch(mp.Rectangle((he_x+0.12, he_y+0.20), band_w, he_h-0.40,
                              fc="#FBE5E5", ec="#C0504D", lw=0.8))
    ax.add_patch(mp.Rectangle((he_x+he_w-0.12-band_w, he_y+0.20), band_w, he_h-0.40,
                              fc="#E1ECF7", ec="#1F4E79", lw=0.8))
    for i in range(13):
        ystr = he_y + 0.45 + i*(he_h-0.90)/12
        ax.plot([he_x+0.12+band_w+0.05, he_x+he_w-0.12-band_w-0.05],
                [ystr, ystr], "-", color="#888888", lw=0.5)
    ax.text(he_x+he_w/2, he_y-0.30, f"HE  ({R['HE_recommended']})",
            ha="center", va="top", fontsize=11, fontweight="bold")
    y_top_port = he_y + he_h - 0.70
    y_bot_port = he_y + 0.70

    # Brine loop (left)
    ax.annotate("", xy=(he_x, y_top_port), xytext=(0.5, y_top_port),
                arrowprops=dict(arrowstyle="-|>,head_width=0.4,head_length=0.5",
                                color="#A52A2A", lw=3.5))
    ax.annotate("", xy=(0.5, y_bot_port), xytext=(he_x, y_bot_port),
                arrowprops=dict(arrowstyle="-|>,head_width=0.4,head_length=0.5",
                                color="#5B1A1A", lw=3.5))
    ax.plot([0.5, 0.5], [y_top_port, y_bot_port],
            color="#888888", lw=1.0, linestyle="--")
    ax.text(0.6, y_top_port+0.35, f"T_brine_in {R['T_brine_in']:.2f} °C",
            fontsize=12, color="#A52A2A", fontweight="bold")
    ax.text(0.6, y_bot_port-0.50, f"T_brine_out {R['T_brine_out']:.2f} °C",
            fontsize=12, color="#5B1A1A", fontweight="bold")
    ax.text(2.0, (y_top_port+y_bot_port)/2,
            f"Flow {R['Q_brine']:.2f} L/s\n  = {R['m_brine']:.2f} kg/s",
            fontsize=11, ha="center", color="#A52A2A")

    # Working-fluid loop (right)
    x_he_r = he_x + he_w
    ax.annotate("", xy=(7.0, y_top_port), xytext=(x_he_r, y_top_port),
                arrowprops=dict(arrowstyle="-|>,head_width=0.4,head_length=0.5",
                                color="#1F4E79", lw=3.5))
    ax.text(x_he_r+0.1, y_top_port+0.35,
            f"T_HE_out {R['T_HE_out']:.2f} °C",
            fontsize=11, color="#1F4E79", fontweight="bold")
    ax.annotate("", xy=(17.0, y_top_port), xytext=(7.0, y_top_port),
                arrowprops=dict(arrowstyle="-|>,head_width=0.4,head_length=0.5",
                                color="#2E7D32", lw=3.5))
    ax.text(12.0, y_top_port+0.35,
            f"Supply pipeline {L_pipe:.0f} m   "
            f"(ΔT = {R['T_HE_out']-R['T_DHS_in']:.2f} K, "
            f"q' ≈ {R['Q_loss_sup']/L_pipe:.1f} W/m)",
            ha="center", fontsize=10, color="#2E7D32")
    ax.annotate("", xy=(17.0, 5.6), xytext=(17.0, y_top_port),
                arrowprops=dict(arrowstyle="-|>,head_width=0.4,head_length=0.5",
                                color="#7B2CBF", lw=3.5))
    ax.text(17.25, 6.0, f"T_DHS_in {R['T_DHS_in']:.2f} °C",
            fontsize=11, color="#7B2CBF", fontweight="bold")

    # DH station box
    ax.add_patch(mp.Rectangle((15.5, 3.4), 3.0, 2.2,
                              fc="#FFFFFF", ec="black", lw=1.8))
    ax.text(17.0, 4.50, "DH station", ha="center", va="center",
            fontsize=12, fontweight="bold")

    ax.annotate("", xy=(17.0, y_bot_port), xytext=(17.0, 3.4),
                arrowprops=dict(arrowstyle="-|>,head_width=0.4,head_length=0.5",
                                color="#7B2CBF", lw=3.5))
    ax.text(17.25, 3.0, f"T_DHS_out {R['T_DHS_out']:.2f} °C",
            fontsize=11, color="#7B2CBF", fontweight="bold")
    ax.annotate("", xy=(7.0, y_bot_port), xytext=(17.0, y_bot_port),
                arrowprops=dict(arrowstyle="-|>,head_width=0.4,head_length=0.5",
                                color="#2E7D32", lw=3.5))
    ax.text(12.0, y_bot_port-0.55,
            f"Return pipeline {L_pipe:.0f} m   "
            f"(ΔT = {R['T_DHS_out']-R['T_HE_in']:.2f} K, "
            f"q' ≈ {R['Q_loss_ret']/L_pipe:.1f} W/m)",
            ha="center", fontsize=10, color="#2E7D32")
    ax.annotate("", xy=(x_he_r, y_bot_port), xytext=(7.0, y_bot_port),
                arrowprops=dict(arrowstyle="-|>,head_width=0.4,head_length=0.5",
                                color="#1F4E79", lw=3.5))
    ax.text(x_he_r+0.1, y_bot_port-0.50,
            f"T_HE_in {R['T_HE_in']:.2f} °C",
            fontsize=11, color="#1F4E79", fontweight="bold")
    ax.text(8.5, (y_top_port+y_bot_port)/2,
            f"Flow {R['m_water']:.2f} kg/s\n  = {R['Q_vol']*1000:.2f} L/s",
            ha="center", fontsize=11, color="#1F4E79")

    # Footer
    summary = (f"Q_HE = {P['Q_W']/1e6:.2f} MW   |   "
               f"Pump P_elec = {R['P_elec_W']/1000:.1f} kW   |   "
               f"HE chosen: {R['HE_recommended']} ({P['plate_class']} plates), "
               f"A = {R['HE_chosen']['A_total_m2']:.0f} m²"
               + ("" if P["fits_single_frame"] else
                  f" / {P['N_frames_needed']} parallel frames")
               + f"   |   LMTD = {P['LMTD_K']:.2f} K   |   "
               f"U = {P['U_Wm2K']:.0f} W/(m²·K)   |   "
               f"ε = {P['effectiveness']*100:.1f} %   |   "
               f"approach = {R['approach_hot']:.1f}/{R['approach_cold']:.1f} K")
    ax.text(11, 0.3, summary, ha="center", fontsize=9.5,
            bbox=dict(boxstyle="round,pad=0.4", fc="#F2F2F2", ec="black"))

    plt.tight_layout()
    path = os.path.join(save_dir, f"system_scheme_{R['option']}.png")
    plt.savefig(path, dpi=200, bbox_inches="tight"); plt.close(fig)
    return path


# ==============================================================================
# PART 16 — SENSITIVITY SWEEP PLOT (2×2)
# ==============================================================================
def draw_sensitivity_plot(results_at_operating_points, save_dir):
    """Sweep T_DHS_out from SWEEP_MIN to SWEEP_MAX and plot Q, m_water,
    A_PHE, OPEX as functions of T_DHS_out, with chosen operating points
    overlaid."""
    import matplotlib.pyplot as plt

    Ts = []
    Qs, mws, As, OPs = [], [], [], []
    failed = []
    t = SWEEP_T_DHS_OUT_MIN
    while t <= SWEEP_T_DHS_OUT_MAX + 1e-9:
        try:
            # Use auto-approaches AND auto plate class
            R = run_with_fallback("sweep", t, None, None,
                                  plate_class_in="auto", verbose=False)
            if R is not None:
                Ts.append(t)
                Qs.append(R["PHE"]["Q_W"]/1e6)
                mws.append(R["m_water"])
                As.append(R["PHE"]["A_total_m2"])
                OPs.append(R["OPEX_excl"])
            else:
                failed.append(t)
        except Exception:
            failed.append(t)
        t += SWEEP_T_DHS_OUT_STEP

    fig, axs = plt.subplots(2, 2, figsize=(12, 8))
    fig.suptitle(f"Parametric sensitivity sweep — T_DHS_out from "
                 f"{SWEEP_T_DHS_OUT_MIN:.0f} to {SWEEP_T_DHS_OUT_MAX:.0f} °C",
                 fontsize=13, fontweight="bold", y=0.995)

    panels = [
        (axs[0,0], Qs,  "Q_HE [MW]",                "#1F4E79", "Q_W", 1e-6),
        (axs[0,1], mws, "ṁ_water [kg/s]",           "#2E7D32", "m_water", 1.0),
        (axs[1,0], As,  "PHE area [m²]",            "#C0504D", "A_total_m2", 1.0),
        (axs[1,1], OPs, "OPEX excl. VAT [EUR/yr]",  "#7B2CBF", "OPEX_excl", 1.0),
    ]
    op_colors = {"A":"#1F4E79", "B":"#2E7D32", "C":"#C0504D"}
    for ax, ys, ylabel, color, key, scale in panels:
        ax.plot(Ts, ys, "-", color=color, lw=2.0, marker="o", markersize=4,
                label="auto-approach sweep")
        # Overlay operating points
        for R in results_at_operating_points:
            if R is None: continue
            if key == "Q_W":         y_pt = R["PHE"]["Q_W"]*scale
            elif key == "A_total_m2":y_pt = R["PHE"]["A_total_m2"]*scale
            else:                    y_pt = R[key]*scale
            ax.plot(R["T_DHS_out"], y_pt, "s",
                    color=op_colors.get(R["option"], "black"),
                    markersize=12, markeredgecolor="black", markeredgewidth=1.5,
                    label=f"Option {R['option']}", zorder=10)
            ax.annotate(R["option"], xy=(R["T_DHS_out"], y_pt),
                        xytext=(8, 8), textcoords="offset points",
                        fontsize=11, fontweight="bold",
                        color=op_colors.get(R["option"], "black"))
        ax.set_xlabel("T_DHS_out [°C]", fontsize=11)
        ax.set_ylabel(ylabel, fontsize=11)
        ax.grid(True, alpha=0.4)
        ax.set_xlim(SWEEP_T_DHS_OUT_MIN-2, SWEEP_T_DHS_OUT_MAX+2)
        # Format y-tick numbers nicely
        if key == "OPEX_excl":
            ax.yaxis.set_major_formatter(plt.matplotlib.ticker.FuncFormatter(
                lambda x, _: f"{x/1000:.1f}k" if x >= 1000 else f"{x:.0f}"))

    # legend on first panel only
    axs[0,0].legend(loc="best", fontsize=9, framealpha=0.9)

    plt.tight_layout(rect=[0,0,1,0.97])
    path = os.path.join(save_dir, "sensitivity_plot.png")
    plt.savefig(path, dpi=200, bbox_inches="tight"); plt.close(fig)
    return path, failed


# ==============================================================================
# PART 17 — ENTRY POINT
# ==============================================================================
def _is_skip(v):
    """A T_DHS_out value of None or 0 means 'skip this option'."""
    if v is None: return True
    try:
        return float(v) <= 0
    except (TypeError, ValueError):
        return True


if __name__ == "__main__":
    options = [
        ("A", T_DHS_out_A, approach_hot_A, approach_cold_A, PHE_plate_class_A),
        ("B", T_DHS_out_B, approach_hot_B, approach_cold_B, PHE_plate_class_B),
        ("C", T_DHS_out_C, approach_hot_C, approach_cold_C, PHE_plate_class_C),
    ]

    results = []
    if VERBOSE:
        print()
        print("="*78)
        print(" GEOTHERMAL DH — SENSITIVITY ANALYSIS (Options A / B / C)")
        print("="*78)
        print(f" Common inputs:  T_brine_in = {T_brine_in:.1f} °C   "
              f"Q_brine = {Q_brine:.1f} L/s   L_pipe = {L_pipe:.0f} m   "
              f"FLH = {FLH:.0f} h/yr")
        print(f" Tariff = {ELEC_TARIFF_EUR_kWh_excl_VAT:.4f} EUR/kWh excl. VAT, "
              f"VAT = {VAT_RATE*100:.0f}%")
        print(f" T_soil_january = {T_soil_january:.1f} °C   "
              f"burial = {burial_depth:.1f} m")
        print()

    for label, T_out, a_hot, a_cold, p_cls in options:
        if _is_skip(T_out):
            if VERBOSE:
                print(f"  Option {label}: SKIPPED (T_DHS_out_{label} = "
                      f"{T_out!r}; set a positive number to enable)")
            results.append(None)
            continue
        try:
            R = run_with_fallback(label, float(T_out), a_hot, a_cold,
                                  plate_class_in=p_cls, verbose=False)
            results.append(R)
            if VERBOSE: print_case(R)
        except ValueError as e:
            print(f"\n  Option {label}: CALCULATION FAILED — {e}\n")
            results.append(None)

    if VERBOSE and sum(r is not None for r in results) > 1:
        print_comparison(results)

    save_dir = (os.path.dirname(os.path.abspath(__file__))
                if "__file__" in dir() else os.getcwd())

    if SAVE_EXCEL:
        try:
            p = write_excel(results, save_dir)
            if VERBOSE: print(f"\nExcel saved             → {p}")
        except ImportError:
            if VERBOSE: print("(openpyxl not installed → no Excel)")

    if SAVE_PHE_SCHEMES:
        try:
            for R in results:
                if R is None: continue
                p = draw_PHE_scheme(R, save_dir)
                if VERBOSE: print(f"PHE schematic (Opt {R['option']}) → {p}")
        except ImportError:
            if VERBOSE: print("(matplotlib not installed → no PHE schematics)")

    if SAVE_SYSTEM_SCHEMES:
        try:
            for R in results:
                if R is None: continue
                p = draw_system_scheme(R, save_dir)
                if VERBOSE: print(f"System schematic (Opt {R['option']}) → {p}")
        except ImportError:
            if VERBOSE: print("(matplotlib not installed → no system schematics)")

    if SAVE_SENSITIVITY_PLOT:
        try:
            p, failed = draw_sensitivity_plot(results, save_dir)
            if VERBOSE:
                print(f"Sensitivity plot         → {p}")
                if failed:
                    print(f"  (sweep points that failed: {failed})")
        except ImportError:
            if VERBOSE: print("(matplotlib not installed → no sensitivity plot)")
